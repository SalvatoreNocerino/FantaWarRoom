# -*- coding: utf-8 -*-
"""
Scrive nel foglio Simulazioni i risultati del confronto A-F a componenti
incrementali (1000 aste x strategia, la mia squadra usa la strategia in
test, le altre 7 usano il motore completo F come avversarie fisse - isola
l'effetto causale della mia scelta). NON tocca le formule di pricing del
motore live (Output Utente, Parametri Modello, ecc.) come richiesto.
"""
import json
import openpyxl

PATH = r"C:\Users\salva\Projects\FantaWarRoom\Excel_Only\Strategia_Asta_FINAL_GPT.xlsx"

with open(r"C:\Users\salva\Projects\FantaWarRoom\Excel_Only\_sim4\full_comparison_results.json") as f:
    R = json.load(f)

wb = openpyxl.load_workbook(PATH)
sim = wb["Simulazioni"]

# rimuove eventuali celle unite pre-esistenti nell'intervallo che riscrivo
for rng in list(sim.merged_cells.ranges):
    if rng.min_row <= 20:
        sim.unmerge_cells(str(rng))

sim["A1"] = "SIMULAZIONI — confronto strategie A-F (1.000 aste x strategia, componenti incrementali)"
sim["A2"] = (
    "Metodologia: la MIA squadra (Team1) usa la strategia in test; le altre 7 usano sempre il motore "
    "completo F (avversarie fisse) — cosi' isoliamo l'effetto causale di CAMBIARE la mia strategia, a "
    "parita' di comportamento avversario. 1.000 aste indipendenti per strategia (seed 20000-20999), 8 "
    "squadre, budget 1000, rosa 3P-8D-8C-6A, listone reale 491 giocatori. Il meccanismo 'brucia budget' di "
    "fine asta e' costante su tutte le varianti (non e' una leva testata, e' una dinamica dell'asta reale)."
)

headers = ["Strategia", "Componenti Attivi", "FVM Medio", "FVM Mediana", "Deviazione Standard",
           "% Vittorie (FVM piu' alto delle 8)", "% Rose Complete (25/25)", "Crediti Inutilizzati (medio)",
           "FVM per Credito Speso", "Note"]
for i, h in enumerate(headers):
    sim.cell(row=4, column=1 + i, value=h)

componenti = {
    "A": "FVM puro (nessun correttivo)",
    "B": "+ Performance storica (Fantamedia reale 2025/26 vs media ruolo)",
    "C": "+ Scarsita' di ruolo (liquidita' di mercato)",
    "D": "+ Quality Scarcity (fasce TOP/BUONO/MEDIA/SCOMMESSA)",
    "E": "+ Auction State (titolari/riserve pianificati)",
    "F": "+ Competition Pressure (bisogno + budget avversari)",
}
note = {
    "A": "Baseline. Senza altre squadre 'furbe' a fare da riferimento, e' il punto di partenza.",
    "B": (
        "Effetto quasi nullo (+0.4% FVM vs A). Il fattore performance e' limitato a +/-25% e la "
        "Fantamedia e' gia' abbastanza correlata al FVM di mercato: aggiunge poco segnale extra."
    ),
    "C": (
        "IL SALTO PIU' GRANDE: +11.2% FVM vs B (930->1034), win% quasi raddoppia (27%->39%). La "
        "scarsita' di ruolo (liberi/slot residui) e' la leva con l'impatto maggiore, di gran lunga."
    ),
    "D": (
        "Effetto piatto o leggermente negativo vs C (-0.3% FVM). Il premio per giocatori TOP/BUONO in "
        "fascia scarsa non produce un beneficio misurabile aggiuntivo rispetto alla sola scarsita' di "
        "mercato (C) in questa simulazione."
    ),
    "E": (
        "+2.4% FVM vs D (1031->1056), ma introduce un rischio strutturale: la classifica "
        "titolare/riserva si basa su QUANTI ne hai gia' comprati (conteggio), non sulla qualita' del "
        "giocatore valutato ora. Una squadra che riempie la quota titolari con giocatori mediocri "
        "sottovaluta sistematicamente un giocatore forte dello stesso ruolo arrivato dopo (verificato: "
        "Dimarco, FVM 265, venduto a 51 crediti in una sessione con tutte le squadre su F)."
    ),
    "F": (
        "PEGGIORA NETTAMENTE (-23.5% FVM vs E, 1056->807; win% crolla al 12.2%, cioe' il livello di "
        "una squadra media senza alcun vantaggio, 1/8=12.5%). Causa verificata: quando TUTTE le squadre "
        "usano Competition Pressure, il moltiplicatore (fino a 2.5x) si applica anche agli SLOT RISERVA "
        "poco pregiati che quasi tutte le squadre devono ancora riempire in tarda asta, scatenando aste "
        "competitive su giocatori quasi inutili. Verificato su 100 aste: sotto F il 57.9% del budget "
        "finisce su giocatori MEDIA/SCOMMESSA (contro 47.4% con A puro), e il prezzo medio pagato per un "
        "giocatore TOP/BUONO SCENDE (45.7 crediti) invece di salire (69.0 con A) — il budget viene "
        "eroso sui riempitivi prima di arrivare ai giocatori che contano. NON e' un artefatto: la logica "
        "di combinare moltiplicativamente 4 fattori (performance x scarsita' x quality x competition) "
        "amplifica troppo i segnali quando piu' condizioni di 'scarsita'' scattano insieme sugli stessi "
        "slot. Da NON portare cosi' com'e' nel motore live."
    ),
}

for i, s in enumerate(("A", "B", "C", "D", "E", "F")):
    r = 5 + i
    d = R[s]
    sim.cell(row=r, column=1, value=f"Strategia {s}")
    sim.cell(row=r, column=2, value=componenti[s])
    sim.cell(row=r, column=3, value=round(d["fvm_medio"], 1))
    sim.cell(row=r, column=4, value=round(d["fvm_mediana"], 1))
    sim.cell(row=r, column=5, value=round(d["fvm_sd"], 1))
    sim.cell(row=r, column=6, value=round(d["win_pct"], 1))
    sim.cell(row=r, column=7, value=round(d["complete_pct"], 1))
    sim.cell(row=r, column=8, value=round(d["resid_medio"], 1))
    sim.cell(row=r, column=9, value=round(d["fvm_per_credito"], 3))
    sim.cell(row=r, column=10, value=note[s])

sim["A12"] = "CONCLUSIONE (per non aggiungere formule sulla base di intuizioni, come richiesto)"
sim["A13"] = (
    "Componenti che migliorano REALMENTE la rosa (misurato): Scarsita' di ruolo (C) — netto, il piu' "
    "importante. Auction State (E) — positivo ma con un limite strutturale da correggere (classificazione "
    "titolare/riserva per conteggio, non per qualita' del giocatore valutato)."
)
sim["A14"] = (
    "Componenti che NON hanno mostrato beneficio misurabile in questa simulazione: Performance storica "
    "(B) — effetto trascurabile. Quality Scarcity (D) — piatto sopra la scarsita' di mercato (C)."
)
sim["A15"] = (
    "Componente che PEGGIORA il risultato cosi' com'e' implementata: Competition Pressure (F) — "
    "causa verificata (bidding war su slot riserva a basso valore, non su titolari). Non integrarlo nel "
    "motore live senza prima rivedere come si combina con Auction State e senza smorzare l'effetto "
    "moltiplicativo quando piu' segnali di scarsita' scattano insieme."
)
sim["A16"] = "Nessuna modifica e' stata applicata al motore live (Output Utente, Parametri Modello) in questo passaggio, come richiesto."

wb.save(PATH)
print("Simulazioni aggiornato.")
