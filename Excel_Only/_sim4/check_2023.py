import openpyxl, io
wb = openpyxl.load_workbook(r"C:\Users\salva\Projects\FantaWarRoom\Excel_Only\Dati\Rose_fantalba-2023.xlsx", data_only=True)
ws = wb["TutteLeRose"]
out = io.open(r"C:\Users\salva\Projects\FantaWarRoom\Excel_Only\_sim4\dump_2023_full.txt", "w", encoding="utf-8")
out.write("dim %s\n" % ws.dimensions)
for r in range(1, 10):
    vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
    out.write(f"row{r}: {vals}\n")
out.close()
print("done")
