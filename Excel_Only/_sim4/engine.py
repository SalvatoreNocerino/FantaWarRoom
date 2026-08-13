# -*- coding: utf-8 -*-
"""
Motore unico a flag incrementali per confrontare Strategia A-F come richiesto:
A = FVM puro
B = A + Performance storica (Fantamedia reale 2025/26 vs media ruolo)
C = B + Scarsita' di ruolo (liquidita' di mercato: liberi/slot residui lega)
D = C + Quality Scarcity (fasce TOP/BUONO/MEDIA/SCOMMESSA per percentile FVM)
E = D + Auction State (pianificazione budget a due livelli titolari/riserve)
F = E + Competition Pressure (quante squadre rivali competono ora per quel tier)

Il meccanismo "brucia budget" di fine asta (i crediti non spesi sono persi)
resta COSTANTE su tutte le varianti: non e' una strategia, e' una dinamica
dell'asta reale (validata nella sessione precedente). Isolarlo cosi' evita di
confondere l'effetto delle singole leve di pricing con differenze di spesa
totale.
"""
import random
import openpyxl

SRC = r"C:\Users\salva\Projects\FantaWarRoom\Excel_Only\Strategia_Asta_FINAL_GPT.xlsx"
N_TEAMS = 8
BUDGET = 1000
ROSTER = {"P": 3, "D": 8, "C": 8, "A": 6}
TITOLARI = {"P": 1, "D": 3, "C": 4, "A": 3}
ROLE_BUDGET_SHARE = {"P": 0.093, "D": 0.116, "C": 0.278, "A": 0.515}
TITOLARI_BUDGET_SHARE = {"P": 0.88, "D": 0.85, "C": 0.95, "A": 0.93}
PRESENZE_MIN = 10
PESO_PERFORMANCE = 0.2
PESO_SCARSITA = 0.3
PERC_TOP, PERC_BUONO, PERC_MEDIA = 0.9, 0.7, 0.4

STRATEGIE = {
    "A": dict(performance=False, market_scarcity=False, quality_tiers=False, auction_state=False, competition_pressure=False),
    "B": dict(performance=True, market_scarcity=False, quality_tiers=False, auction_state=False, competition_pressure=False),
    "C": dict(performance=True, market_scarcity=True, quality_tiers=False, auction_state=False, competition_pressure=False),
    "D": dict(performance=True, market_scarcity=True, quality_tiers=True, auction_state=False, competition_pressure=False),
    "E": dict(performance=True, market_scarcity=True, quality_tiers=True, auction_state=True, competition_pressure=False),
    "F": dict(performance=True, market_scarcity=True, quality_tiers=True, auction_state=True, competition_pressure=True),
}


def load_players():
    wb = openpyxl.load_workbook(SRC, data_only=True)
    ws = wb["Input Giocatori 2026_27"]
    players = []
    for r in range(3, ws.max_row + 1):
        pid = ws.cell(row=r, column=1).value
        ruolo = ws.cell(row=r, column=2).value
        nome = ws.cell(row=r, column=3).value
        fvm = ws.cell(row=r, column=7).value
        if pid and ruolo in ROSTER and nome:
            players.append({"id": pid, "ruolo": ruolo, "nome": nome, "fvm": fvm or 0})

    stat_ws = wb["Input Statistiche 2025_26"]
    stats = {}
    for r in range(3, stat_ws.max_row + 1):
        pid = stat_ws.cell(row=r, column=1).value
        pv = stat_ws.cell(row=r, column=6).value
        fm = stat_ws.cell(row=r, column=8).value
        if pid is not None:
            stats[pid] = (pv, fm)

    role_fm_sum = {r: 0.0 for r in ROSTER}
    role_fm_n = {r: 0 for r in ROSTER}
    for p in players:
        pv, fm = stats.get(p["id"], (None, None))
        if pv is not None and fm is not None and pv >= PRESENZE_MIN:
            role_fm_sum[p["ruolo"]] += fm
            role_fm_n[p["ruolo"]] += 1
    role_fm_avg = {r: (role_fm_sum[r] / role_fm_n[r] if role_fm_n[r] else None) for r in ROSTER}

    for p in players:
        pv, fm = stats.get(p["id"], (None, None))
        avg = role_fm_avg[p["ruolo"]]
        if pv is not None and fm is not None and pv >= PRESENZE_MIN and avg:
            rel = (fm - avg) / avg
            p["perf_mult"] = max(0.75, min(1.25, 1 + PESO_PERFORMANCE * max(-1, min(1, rel))))
        else:
            p["perf_mult"] = 1.0
    return players


def build_tiers(players):
    tiers = {}
    for role in ROSTER:
        role_players = sorted([p for p in players if p["ruolo"] == role], key=lambda p: -p["fvm"])
        n_role = len(role_players)
        pool_n = N_TEAMS * ROSTER[role]
        tit_n = N_TEAMS * TITOLARI[role]
        titolari_ids = {p["id"] for p in role_players[:tit_n]}
        riserve_ids = {p["id"] for p in role_players[tit_n:pool_n]}
        top_n = max(1, round(n_role * (1 - PERC_TOP)))
        buono_n = max(1, round(n_role * (1 - PERC_BUONO))) - top_n
        media_n = max(1, round(n_role * (1 - PERC_MEDIA))) - top_n - buono_n
        top_ids = {p["id"] for p in role_players[:top_n]}
        buono_ids = {p["id"] for p in role_players[top_n:top_n + max(0, buono_n)]}
        tiers[role] = dict(
            titolari_ids=titolari_ids, riserve_ids=riserve_ids,
            titolari_fvm_avg=(sum(p["fvm"] for p in role_players[:tit_n]) / max(1, tit_n)),
            riserve_fvm_avg=(sum(p["fvm"] for p in role_players[tit_n:pool_n]) / max(1, pool_n - tit_n)) or 1,
            top_buono_ids=top_ids | buono_ids,
            n_role_total=n_role,
        )
    return tiers


def run_auction(players, tiers, seed, my_flags, opp_flags, noise_sigma=0.15, noise_clip=(0.7, 1.4)):
    rng = random.Random(seed)
    teams = {}
    for i in range(N_TEAMS):
        teams[f"Team{i+1}"] = dict(
            budget=BUDGET, roster=[],
            tit_bought={r: 0 for r in ROSTER}, ris_bought={r: 0 for r in ROSTER},
            spent_tit={r: 0 for r in ROSTER}, spent_ris={r: 0 for r in ROSTER},
            flags=(my_flags if i == 0 else opp_flags),
        )
    remaining_fvm = {r: {"titolari": sum(p["fvm"] for p in players if p["id"] in tiers[r]["titolari_ids"]),
                          "riserve": sum(p["fvm"] for p in players if p["id"] in tiers[r]["riserve_ids"])}
                      for r in ROSTER}
    remaining_n = {r: {"titolari": len(tiers[r]["titolari_ids"]), "riserve": len(tiers[r]["riserve_ids"])}
                   for r in ROSTER}
    remaining_fvm_role = {r: sum(p["fvm"] for p in players if p["ruolo"] == r) for r in ROSTER}
    remaining_n_role = {r: sum(1 for p in players if p["ruolo"] == r) for r in ROSTER}
    remaining_topbuono_role = {r: sum(1 for p in players if p["ruolo"] == r and p["id"] in tiers[r]["top_buono_ids"]) for r in ROSTER}

    order = players[:]
    rng.shuffle(order)
    league_slots_left = {r: N_TEAMS * ROSTER[r] for r in ROSTER}
    TOTAL_SLOTS_TEAM = sum(ROSTER.values())
    TOTAL_SLOTS_LEAGUE = TOTAL_SLOTS_TEAM * N_TEAMS
    slots_filled_league = 0

    def player_tier(role, pid):
        if pid in tiers[role]["titolari_ids"]:
            return "titolari"
        if pid in tiers[role]["riserve_ids"]:
            return "riserve"
        return "floor"

    def bid_ceiling(t, role, p, market_tier, flags, pace_league):
        use_tit = t["tit_bought"][role] < TITOLARI[role]
        role_budget = ROLE_BUDGET_SHARE[role] * BUDGET

        if flags["auction_state"]:
            use_tier = "titolari" if (use_tit and market_tier != "floor") else "riserve"
            if use_tier == "titolari":
                target = role_budget * TITOLARI_BUDGET_SHARE[role]
                spent = t["spent_tit"][role]
                slots_needed = TITOLARI[role] - t["tit_bought"][role]
                avg_fvm = (remaining_fvm[role]["titolari"] / remaining_n[role]["titolari"]) if remaining_n[role]["titolari"] > 0 else tiers[role]["titolari_fvm_avg"]
            else:
                target = role_budget * (1 - TITOLARI_BUDGET_SHARE[role])
                spent = t["spent_ris"][role]
                slots_needed = (ROSTER[role] - TITOLARI[role]) - t["ris_bought"][role]
                if market_tier == "floor":
                    avg_fvm = max(1, tiers[role]["riserve_fvm_avg"] * 0.3)
                else:
                    avg_fvm = (remaining_fvm[role]["riserve"] / remaining_n[role]["riserve"]) if remaining_n[role]["riserve"] > 0 else tiers[role]["riserve_fvm_avg"]
            if slots_needed <= 0:
                return None, use_tier
        else:
            use_tier = "titolari" if (use_tit and market_tier != "floor") else "riserve"  # per bookkeeping soltanto
            target = role_budget
            spent = t["spent_tit"][role] + t["spent_ris"][role]
            slots_needed = ROSTER[role] - (t["tit_bought"][role] + t["ris_bought"][role])
            avg_fvm = (remaining_fvm_role[role] / remaining_n_role[role]) if remaining_n_role[role] > 0 else 1
            if slots_needed <= 0:
                return None, use_tier

        remaining_target = max(1, target - spent)
        per_slot = remaining_target / slots_needed
        weight = p["fvm"] / max(1, avg_fvm) if avg_fvm else 1.0
        weight = max(0.15, min(4.0, weight))
        ceiling = per_slot * weight

        if flags["performance"]:
            ceiling *= p["perf_mult"]

        if flags["market_scarcity"]:
            avail_ratio = (remaining_n_role[role] / max(1, league_slots_left[role]))
            mkt_mult = max(0.7, min(1.3, 1 + PESO_SCARSITA * (1 - avail_ratio)))
            ceiling *= mkt_mult

        if flags["quality_tiers"] and p["id"] in tiers[role]["top_buono_ids"]:
            denom = (remaining_n[role]["titolari"] if flags["auction_state"] else max(1, league_slots_left[role]))
            q_ratio = remaining_topbuono_role[role] / max(1, denom)
            q_mult = max(0.7, min(1.5, 1 + 0.3 * (1 - q_ratio)))
            ceiling *= q_mult

        if flags["competition_pressure"]:
            if use_tier == "titolari":
                n_need = sum(1 for tt in teams.values() if tt["tit_bought"][role] < TITOLARI[role])
                pool_left = max(1, remaining_n[role]["titolari"])
            else:
                n_need = sum(1 for tt in teams.values() if (ROSTER[role] - TITOLARI[role]) - tt["ris_bought"][role] > 0)
                pool_left = max(1, remaining_n[role]["riserve"] + 1)
            comp_mult = max(0.6, min(2.5, n_need / pool_left))
            ceiling *= comp_mult

        # brucia budget di fine asta - costante per tutte le strategie
        slots_total_left = (TITOLARI[role] - t["tit_bought"][role]) + ((ROSTER[role] - TITOLARI[role]) - t["ris_bought"][role])
        other_open = sum((TITOLARI[r] - t["tit_bought"][r]) + ((ROSTER[r] - TITOLARI[r]) - t["ris_bought"][r]) for r in ROSTER if r != role)
        own_slots_left_all = slots_total_left + other_open
        pace_factor = 0.35 + 1.15 * pace_league
        endgame_floor = (t["budget"] / max(1, own_slots_left_all)) * pace_factor
        ceiling = max(ceiling, endgame_floor)

        return ceiling, use_tier

    for p in order:
        role = p["ruolo"]
        if league_slots_left[role] <= 0:
            continue
        market_tier = player_tier(role, p["id"])
        pace_league = slots_filled_league / TOTAL_SLOTS_LEAGUE if TOTAL_SLOTS_LEAGUE else 0

        bids, tiers_used = {}, {}
        for tname, t in teams.items():
            slots_total_left = (TITOLARI[role] - t["tit_bought"][role]) + ((ROSTER[role] - TITOLARI[role]) - t["ris_bought"][role])
            if slots_total_left <= 0:
                continue
            other_open = sum((TITOLARI[r] - t["tit_bought"][r]) + ((ROSTER[r] - TITOLARI[r]) - t["ris_bought"][r]) for r in ROSTER if r != role)
            reserve = other_open + max(0, slots_total_left - 1)
            spendable = max(0, t["budget"] - reserve)
            if spendable <= 0:
                continue
            ceiling, use_tier = bid_ceiling(t, role, p, market_tier, t["flags"], pace_league)
            if ceiling is None:
                continue
            noise = min(noise_clip[1], max(noise_clip[0], rng.lognormvariate(0, noise_sigma)))
            bid = min(ceiling * noise, spendable)
            if bid >= 1:
                bids[tname] = bid
                tiers_used[tname] = use_tier
        if not bids:
            continue
        winner = max(bids, key=bids.get)
        sorted_bids = sorted(bids.values(), reverse=True)
        price = int(round(sorted_bids[1])) if len(sorted_bids) > 1 else 1
        price = max(1, min(price, teams[winner]["budget"]))
        wt = teams[winner]
        use_tier = tiers_used[winner]
        wt["budget"] -= price
        if use_tier == "titolari":
            wt["tit_bought"][role] += 1
            wt["spent_tit"][role] += price
        else:
            wt["ris_bought"][role] += 1
            wt["spent_ris"][role] += price
        wt["roster"].append({**p, "prezzo": price, "tier": use_tier})
        league_slots_left[role] -= 1
        slots_filled_league += 1
        if market_tier != "floor":
            remaining_fvm[role][market_tier] -= p["fvm"]
            remaining_n[role][market_tier] -= 1
        remaining_fvm_role[role] -= p["fvm"]
        remaining_n_role[role] -= 1
        if p["id"] in tiers[role]["top_buono_ids"]:
            remaining_topbuono_role[role] -= 1

    return teams


if __name__ == "__main__":
    players = load_players()
    tiers = build_tiers(players)
    print("Giocatori:", len(players))
    for r, t in tiers.items():
        print(r, "n_role_totale=", t["n_role_total"], "top_buono=", len(t["top_buono_ids"]))
