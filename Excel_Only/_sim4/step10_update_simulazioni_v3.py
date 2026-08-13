# -*- coding: utf-8 -*-
import json
import openpyxl

PATH = r"C:\Users\salva\Projects\FantaWarRoom\Excel_Only\Strategia_Asta_FINAL_GPT.xlsx"

with open(r"C:\Users\salva\Projects\FantaWarRoom\Excel_Only\_sim4\full_comparison_results3.json") as f:
    R = json.load(f)

wb = openpyxl.load_workbook(PATH)
sim = wb["Simulazioni"]

for rng in list(sim.merged_cells.ranges):
    if rng.min_row <= 22:
        sim.unmerge_cells(str(rng))
for row in sim.iter_rows(min_row=1, max_row=22):
    for cell in row:
        cell.value = None

sim["A1"] = "SIMULAZIONI — confronto strategie A-F v3 (correzione 4bis applicata, risultato parziale)"
sim["A2"] = (
    "v3 aggiunge alla v2 la correzione 4bis: l'urgenza di spesa di fine asta pesa meno sulle fasce di "
    "riserva basse (0.25-0.7) rispetto ai titolari (1.0), tornando comunque a piena urgenza su tutte le "
    "fasce quando l'asta si avvicina alla fine (per non perdere il risultato di spesa ~90%+ validato in "
    "precedenza). Risultato: MIGLIORAMENTO PARZIALE, non risolutivo (vedi note sotto)."
)

headers = ["Strategia", "Componenti Attivi", "FVM Medio", "FVM Mediana", "Deviazione Standard",
           "% Vittorie", "% Rose Complete", "Crediti Inutilizzati", "FVM per Credito", "Note"]
for i, h in enumerate(headers):
    sim.cell(row=4, column=1 + i, value=h)

componenti = {
    "A": "FVM puro",
    "B": "+ Performance storica",
    "C": "+ Scarsita' di ruolo",
    "D": "+ Quality Scarcity",
    "E": "+ Auction State (4 fasce + urgenza fine asta pesata per fascia)",
    "F": "+ Competition Pressure (gated sulle riserve)",
}
note = {
    "A": "Invariata nel principio.",
    "B": "Effetto marginale confermato ancora una volta.",
    "C": "Ancora identica a B (vedi nota v2: la scarsita' di ruolo come definita raramente scende sotto 1).",
    "D": "Sostanzialmente invariata vs C.",
    "E": (
        "MIGLIORATA rispetto alla v2 (767 -> 807 FVM medio, +5%) ma resta ben sotto A-D (963-978) e "
        "sotto il risultato v1 originale (1056, con la struttura a 2 fasce). La correzione 4bis ha "
        "aiutato ma non risolto: ho trovato la causa. Passare da 2 a 4 fasce ha moltiplicato i punti di "
        "'scatto brusco' (transizione tra fasce basata su QUANTI ne hai comprati, non su quanto e' forte "
        "il giocatore) da 1 (titolari->riserve) a 3 (F1->F2->F3->F4): ogni transizione resta un punto in "
        "cui un giocatore forte arrivato tardi puo' essere valutato con il budget della fascia sbagliata. "
        "Piu' fasce, con questa logica di assegnazione, significa piu' occasioni di errore, non meno."
    ),
    "F": (
        "Sostanzialmente invariata (744 vs 751 in v2). Eredita lo stesso problema di E."
    ),
}

for i, s in enumerate(("A", "B", "C", "D", "E", "F")):
    r = 5 + i
    d = R[s]
    sim.cell(row=r, column=1, value=f"Strategia {s}")
    sim.cell(row=r, column=2, value=componenti[s])
    sim.cell(row=r, column=3, value=round(d["fvm_medio"], 1))
    sim.cell(row=r, column=4, value=round(d["fvm_mediana"], 1))
    sim.cell(row=r, column=5, value=round(d["fvm_sd"], 1))
    sim.cell(row=r, column=6, value=round(d["win_pct"], 1))
    sim.cell(row=r, column=7, value=round(d["complete_pct"], 1))
    sim.cell(row=r, column=8, value=round(d["resid_medio"], 1))
    sim.cell(row=r, column=9, value=round(d["fvm_per_credito"], 3))
    sim.cell(row=r, column=10, value=note[s])

sim["A12"] = "STATO: correzione 4bis applicata e verificata (spesa media resta sana, 96%), ma NON risolutiva per E/F."
sim["A13"] = (
    "Diagnosi finale: il problema non e' (solo) l'urgenza di fine asta indifferente alla fascia (corretto "
    "in 4bis) - e' che la fascia di un giocatore viene decisa dalla squadra in base a QUANTI acquisti ha "
    "gia' fatto in quel ruolo, non da QUANTO E' FORTE il giocatore chiamato in quel momento. Con 4 fasce "
    "invece di 2, questo genera 3 punti di transizione invece di 1, quindi piu' occasioni di valutare un "
    "giocatore forte con il budget sbagliato (piu' basso) o uno debole con il budget sbagliato (piu' alto)."
)
sim["A14"] = (
    "Prossimo passo suggerito (non ancora implementato, da confermare): la fascia di un giocatore "
    "dovrebbe dipendere anche da QUANTO E' FORTE lui stesso (il suo rank di mercato), non solo dal "
    "conteggio degli acquisti della squadra - es. una squadra che ha gia' i titolari puo' comunque "
    "valutare un giocatore di rank 'titolare' con un budget intermedio se e' nettamente piu' forte delle "
    "riserve rimaste, invece di scontarlo automaticamente al budget piu' basso disponibile."
)
sim["A15"] = "Nessuna modifica alle formule del motore live in questo passaggio."

wb.save(PATH)
print("Simulazioni v3 aggiornato.")
