# -*- coding: utf-8 -*-
"""
Motore v3: come engine2.py, ma con la correzione 4bis: il meccanismo "non
sprecare crediti" di fine asta ora pesa MENO sulle fasce di riserva basse
rispetto ai titolari, cosi' da non bruciare budget su slot di scarso valore
quando potrebbero ancora arrivare giocatori forti. Il peso torna pieno (1.0)
su tutte le fasce quando l'asta si avvicina alla fine (pace_league -> 1),
per continuare a garantire che i crediti vengano comunque spesi entro la
fine (validato in una sessione precedente: senza questo, la spesa media
crolla al 60% circa).
"""
import random, math
import openpyxl

SRC = r"C:\Users\salva\Projects\FantaWarRoom\Excel_Only\Strategia_Asta_FINAL_GPT.xlsx"
N_TEAMS = 8
BUDGET = 1000
ROSTER = {"P": 3, "D": 8, "C": 8, "A": 6}
TITOLARI = {"P": 1, "D": 3, "C": 4, "A": 3}
ROLE_BUDGET_SHARE = {"P": 0.093, "D": 0.116, "C": 0.278, "A": 0.515}
TITOLARI_BUDGET_SHARE = {"P": 0.88, "D": 0.85, "C": 0.95, "A": 0.93}
RISERVE_TIER_SPLIT = [0.5, 0.3, 0.2]
COMP_PRESSURE_RISERVA_SOGLIA = 1.3

# CORREZIONE 4bis: peso dell'urgenza di fine asta per fascia (1.0 = piena
# urgenza come prima, valori piu' bassi = meno pressione a spendere su quello
# slot specifico finche' l'asta non e' quasi finita).
FASCIA_URGENZA_BASE = {
    "F1_titolari": 1.0,
    "F2_riserve_alte": 0.7,
    "F3_riserve_medie": 0.45,
    "F4_riserve_basse": 0.25,
}

PRESENZE_MIN = 10
PESO_PERFORMANCE = 0.2
PESO_SCARSITA = 0.3
PERC_TOP, PERC_BUONO, PERC_MEDIA = 0.9, 0.7, 0.4

STRATEGIE = {
    "A": dict(performance=False, market_scarcity=False, quality_tiers=False, auction_state=False, competition_pressure=False),
    "B": dict(performance=True, market_scarcity=False, quality_tiers=False, auction_state=False, competition_pressure=False),
    "C": dict(performance=True, market_scarcity=True, quality_tiers=False, auction_state=False, competition_pressure=False),
    "D": dict(performance=True, market_scarcity=True, quality_tiers=True, auction_state=False, competition_pressure=False),
    "E": dict(performance=True, market_scarcity=True, quality_tiers=True, auction_state=True, competition_pressure=False),
    "F": dict(performance=True, market_scarcity=True, quality_tiers=True, auction_state=True, competition_pressure=True),
}


def load_players():
    wb = openpyxl.load_workbook(SRC, data_only=True)
    ws = wb["Input Giocatori 2026_27"]
    players = []
    for r in range(3, ws.max_row + 1):
        pid = ws.cell(row=r, column=1).value
        ruolo = ws.cell(row=r, column=2).value
        nome = ws.cell(row=r, column=3).value
        fvm = ws.cell(row=r, column=7).value
        if pid and ruolo in ROSTER and nome:
            players.append({"id": pid, "ruolo": ruolo, "nome": nome, "fvm": fvm or 0})

    stat_ws = wb["Input Statistiche 2025_26"]
    stats = {}
    for r in range(3, stat_ws.max_row + 1):
        pid = stat_ws.cell(row=r, column=1).value
        pv = stat_ws.cell(row=r, column=6).value
        fm = stat_ws.cell(row=r, column=8).value
        if pid is not None:
            stats[pid] = (pv, fm)

    role_fm_sum = {r: 0.0 for r in ROSTER}
    role_fm_n = {r: 0 for r in ROSTER}
    for p in players:
        pv, fm = stats.get(p["id"], (None, None))
        if pv is not None and fm is not None and pv >= PRESENZE_MIN:
            role_fm_sum[p["ruolo"]] += fm
            role_fm_n[p["ruolo"]] += 1
    role_fm_avg = {r: (role_fm_sum[r] / role_fm_n[r] if role_fm_n[r] else None) for r in ROSTER}

    for p in players:
        pv, fm = stats.get(p["id"], (None, None))
        avg = role_fm_avg[p["ruolo"]]
        if pv is not None and fm is not None and pv >= PRESENZE_MIN and avg:
            rel = (fm - avg) / avg
            p["perf_mult"] = max(0.75, min(1.25, 1 + PESO_PERFORMANCE * max(-1, min(1, rel))))
        else:
            p["perf_mult"] = 1.0
    return players


def build_fasce(role):
    tit_n = TITOLARI[role]
    riserve_n = ROSTER[role] - tit_n
    f2 = max(0, min(riserve_n, math.ceil(riserve_n / 3))) if riserve_n > 0 else 0
    rest = riserve_n - f2
    f3 = max(0, min(rest, math.ceil(rest / 2))) if rest > 0 else 0
    f4 = rest - f3
    counts = [tit_n, f2, f3, f4]
    tit_share = TITOLARI_BUDGET_SHARE[role]
    riserve_share = 1 - tit_share
    shares = [tit_share]
    for w in RISERVE_TIER_SPLIT:
        shares.append(riserve_share * w)
    for i in range(1, 4):
        if counts[i] == 0 and i < 3:
            shares[i + 1] += shares[i]
            shares[i] = 0
    names = ["F1_titolari", "F2_riserve_alte", "F3_riserve_medie", "F4_riserve_basse"]
    return [(names[i], counts[i], shares[i]) for i in range(4) if counts[i] > 0]


def build_tiers(players):
    tiers = {}
    for role in ROSTER:
        role_players = sorted([p for p in players if p["ruolo"] == role], key=lambda p: -p["fvm"])
        n_role = len(role_players)
        fasce = build_fasce(role)
        fascia_ids = {}
        cum = 0
        for name, count, share in fasce:
            n_pool = N_TEAMS * count
            ids = {p["id"] for p in role_players[cum:cum + n_pool]}
            fvm_avg = (sum(p["fvm"] for p in role_players[cum:cum + n_pool]) / max(1, len(ids))) if ids else 1
            fascia_ids[name] = dict(ids=ids, count=count, share=share, fvm_avg=fvm_avg or 1)
            cum += n_pool

        pool_n = cum
        top_n = max(1, round(n_role * (1 - PERC_TOP)))
        buono_n = max(1, round(n_role * (1 - PERC_BUONO))) - top_n
        top_ids = {p["id"] for p in role_players[:top_n]}
        buono_ids = {p["id"] for p in role_players[top_n:top_n + max(0, buono_n)]}

        tiers[role] = dict(
            fasce=fasce, fascia_ids=fascia_ids, pool_n=pool_n,
            top_buono_ids=top_ids | buono_ids, n_role_total=n_role,
        )
    return tiers


def player_fascia(tiers, role, pid):
    for name, d in tiers[role]["fascia_ids"].items():
        if pid in d["ids"]:
            return name
    return "floor"


def run_auction(players, tiers, seed, my_flags, opp_flags, noise_sigma=0.15, noise_clip=(0.7, 1.4)):
    rng = random.Random(seed)
    fascia_names = {role: [f[0] for f in tiers[role]["fasce"]] for role in ROSTER}

    teams = {}
    for i in range(N_TEAMS):
        teams[f"Team{i+1}"] = dict(
            budget=BUDGET, roster=[],
            bought={r: {fn: 0 for fn in fascia_names[r]} for r in ROSTER},
            spent={r: {fn: 0 for fn in fascia_names[r]} for r in ROSTER},
            flags=(my_flags if i == 0 else opp_flags),
        )

    remaining_n = {r: {fn: len(tiers[r]["fascia_ids"][fn]["ids"]) for fn in fascia_names[r]} for r in ROSTER}
    remaining_fvm = {r: {fn: sum(p["fvm"] for p in players if p["id"] in tiers[r]["fascia_ids"][fn]["ids"]) for fn in fascia_names[r]} for r in ROSTER}
    remaining_fvm_role = {r: sum(p["fvm"] for p in players if p["ruolo"] == r) for r in ROSTER}
    remaining_n_role = {r: sum(1 for p in players if p["ruolo"] == r) for r in ROSTER}
    remaining_topbuono_role = {r: sum(1 for p in players if p["ruolo"] == r and p["id"] in tiers[r]["top_buono_ids"]) for r in ROSTER}

    order = players[:]
    rng.shuffle(order)
    league_slots_left = {r: N_TEAMS * ROSTER[r] for r in ROSTER}
    TOTAL_SLOTS_TEAM = sum(ROSTER.values())
    TOTAL_SLOTS_LEAGUE = TOTAL_SLOTS_TEAM * N_TEAMS
    slots_filled_league = 0

    def team_slots_left(t, role):
        return sum(tiers[role]["fasce"][idx][1] - t["bought"][role][fn] for idx, (fn, cnt, sh) in enumerate(tiers[role]["fasce"]))

    def bid_ceiling(t, role, p, market_fascia, flags, pace_league):
        role_budget = ROLE_BUDGET_SHARE[role] * BUDGET
        fasce = tiers[role]["fasce"]

        if flags["auction_state"]:
            use_fascia = None
            for fn, cnt, sh in fasce:
                if t["bought"][role][fn] < cnt:
                    use_fascia = fn
                    break
            if use_fascia is None:
                return None, None
            fn, cnt, sh = next(f for f in fasce if f[0] == use_fascia)
            target = role_budget * sh
            spent = t["spent"][role][use_fascia]
            slots_needed = cnt - t["bought"][role][use_fascia]
            avg_fvm = (remaining_fvm[role][use_fascia] / remaining_n[role][use_fascia]) if remaining_n[role].get(use_fascia, 0) > 0 else tiers[role]["fascia_ids"][use_fascia]["fvm_avg"]
        else:
            use_fascia = fasce[0][0] if market_fascia == "floor" else market_fascia
            target = role_budget
            spent = sum(t["spent"][role].values())
            slots_needed = ROSTER[role] - sum(t["bought"][role].values())
            avg_fvm = (remaining_fvm_role[role] / remaining_n_role[role]) if remaining_n_role[role] > 0 else 1
            if slots_needed <= 0:
                return None, use_fascia

        if slots_needed <= 0:
            return None, use_fascia

        remaining_target = max(1, target - spent)
        per_slot = remaining_target / slots_needed
        weight = p["fvm"] / max(1, avg_fvm) if avg_fvm else 1.0
        weight = max(0.15, min(4.0, weight))
        ceiling = per_slot * weight

        if flags["performance"]:
            ceiling *= p["perf_mult"]

        if flags["market_scarcity"]:
            avail_ratio = (remaining_n_role[role] / max(1, league_slots_left[role]))
            if avail_ratio < 1:
                mkt_mult = min(1.3, 1 + PESO_SCARSITA * (1 - avail_ratio))
                ceiling *= mkt_mult

        if flags["quality_tiers"] and p["id"] in tiers[role]["top_buono_ids"]:
            denom = (remaining_n[role].get("F1_titolari", 1) if flags["auction_state"] else max(1, league_slots_left[role]))
            q_ratio = remaining_topbuono_role[role] / max(1, denom)
            if q_ratio < 1:
                q_mult = min(1.5, 1 + 0.3 * (1 - q_ratio))
                ceiling *= q_mult

        if flags["competition_pressure"]:
            is_titolari = (use_fascia == "F1_titolari")
            apply_pressure = True
            if not is_titolari:
                avg_fascia = (remaining_fvm[role][use_fascia] / remaining_n[role][use_fascia]) if remaining_n[role].get(use_fascia, 0) > 0 else tiers[role]["fascia_ids"].get(use_fascia, {}).get("fvm_avg", 1)
                apply_pressure = p["fvm"] >= COMP_PRESSURE_RISERVA_SOGLIA * max(1, avg_fascia)
            if apply_pressure:
                n_need = sum(1 for tt in teams.values() if tt["bought"][role].get(use_fascia, 0) < next((c for fn, c, s in fasce if fn == use_fascia), 0))
                pool_left = max(1, remaining_n[role].get(use_fascia, 1))
                comp_mult = max(0.6, min(2.5, n_need / pool_left))
                ceiling *= comp_mult

        # CORREZIONE 4bis: urgenza di fine asta pesata per fascia. Vicino alla
        # fine (pace_league -> 1) il peso torna comunque a 1.0 su tutte le
        # fasce, altrimenti il budget rischierebbe di restare inutilizzato.
        slots_total_left = team_slots_left(t, role)
        other_open = sum(team_slots_left(t, r) for r in ROSTER if r != role)
        own_slots_left_all = slots_total_left + other_open
        pace_factor = 0.35 + 1.15 * pace_league
        urgenza_base = FASCIA_URGENZA_BASE.get(use_fascia, 1.0) if flags["auction_state"] else 1.0
        urgenza_effettiva = max(urgenza_base, pace_league)
        endgame_floor = (t["budget"] / max(1, own_slots_left_all)) * pace_factor * urgenza_effettiva
        ceiling = max(ceiling, endgame_floor)

        return ceiling, use_fascia

    for p in order:
        role = p["ruolo"]
        if league_slots_left[role] <= 0:
            continue
        market_fascia = player_fascia(tiers, role, p["id"])
        pace_league = slots_filled_league / TOTAL_SLOTS_LEAGUE if TOTAL_SLOTS_LEAGUE else 0

        bids, fasce_used = {}, {}
        for tname, t in teams.items():
            slots_total_left = team_slots_left(t, role)
            if slots_total_left <= 0:
                continue
            other_open = sum(team_slots_left(t, r) for r in ROSTER if r != role)
            reserve = other_open + max(0, slots_total_left - 1)
            spendable = max(0, t["budget"] - reserve)
            if spendable <= 0:
                continue
            ceiling, use_fascia = bid_ceiling(t, role, p, market_fascia, t["flags"], pace_league)
            if ceiling is None:
                continue
            noise = min(noise_clip[1], max(noise_clip[0], rng.lognormvariate(0, noise_sigma)))
            bid = min(ceiling * noise, spendable)
            if bid >= 1:
                bids[tname] = bid
                fasce_used[tname] = use_fascia
        if not bids:
            continue
        winner = max(bids, key=bids.get)
        sorted_bids = sorted(bids.values(), reverse=True)
        price = int(round(sorted_bids[1])) if len(sorted_bids) > 1 else 1
        price = max(1, min(price, teams[winner]["budget"]))
        wt = teams[winner]
        use_fascia = fasce_used[winner]
        wt["budget"] -= price
        wt["bought"][role][use_fascia] = wt["bought"][role].get(use_fascia, 0) + 1
        wt["spent"][role][use_fascia] = wt["spent"][role].get(use_fascia, 0) + price
        wt["roster"].append({**p, "prezzo": price, "fascia": use_fascia})
        league_slots_left[role] -= 1
        slots_filled_league += 1
        if market_fascia != "floor" and market_fascia in remaining_n[role]:
            remaining_fvm[role][market_fascia] -= p["fvm"]
            remaining_n[role][market_fascia] -= 1
        remaining_fvm_role[role] -= p["fvm"]
        remaining_n_role[role] -= 1
        if p["id"] in tiers[role]["top_buono_ids"]:
            remaining_topbuono_role[role] -= 1

    return teams


if __name__ == "__main__":
    players = load_players()
    tiers = build_tiers(players)
    print("Giocatori:", len(players))
