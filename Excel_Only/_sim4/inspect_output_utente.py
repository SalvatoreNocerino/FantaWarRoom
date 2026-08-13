import openpyxl, io
wb = openpyxl.load_workbook(r"C:\Users\salva\Projects\FantaWarRoom\Excel_Only\_sim4\FINAL_GPT_work.xlsx")
ws = wb["Output Utente"]
out = io.open(r"C:\Users\salva\Projects\FantaWarRoom\Excel_Only\_sim4\dump_output_utente.txt", "w", encoding="utf-8")
out.write("DIM %s\n" % ws.dimensions)
# header row 1
for c in range(1, ws.max_column + 1):
    cell = ws.cell(row=1, column=c)
    out.write("HEADER %s(%s) %s\n" % (cell.coordinate, c, repr(cell.value)))
# row2 formulas (first data row) for every column
for c in range(1, ws.max_column + 1):
    cell = ws.cell(row=2, column=c)
    v = cell.value
    if hasattr(v, "text"):
        v = "ARRAY:" + str(v.text)
    out.write("ROW2 %s %s\n" % (cell.coordinate, repr(v)[:250]))
out.close()
print("done", ws.max_column, ws.max_row)
