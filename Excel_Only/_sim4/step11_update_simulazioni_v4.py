# -*- coding: utf-8 -*-
import json
import openpyxl

PATH = r"C:\Users\salva\Projects\FantaWarRoom\Excel_Only\Strategia_Asta_FINAL_GPT.xlsx"

with open(r"C:\Users\salva\Projects\FantaWarRoom\Excel_Only\_sim4\full_comparison_results4.json") as f:
    R = json.load(f)

wb = openpyxl.load_workbook(PATH)
sim = wb["Simulazioni"]

for rng in list(sim.merged_cells.ranges):
    if rng.min_row <= 24:
        sim.unmerge_cells(str(rng))
for row in sim.iter_rows(min_row=1, max_row=24):
    for cell in row:
        cell.value = None

sim["A1"] = "SIMULAZIONI — confronto strategie A-F v4 (correzione 5: fascia quality-based)"
sim["A2"] = (
    "v4 aggiunge alla v3 la correzione 5: se il giocatore chiamato appartiene, per rank di mercato, a "
    "una fascia migliore di quella che la squadra sta riempiendo per conteggio, e il suo FVM supera del "
    "30% la media della fascia-target, il peso del bid riflette la sua vera fascia di mercato (non solo "
    "quella bassa del conteggio). Risultato: fix implementato e verificato attivo (si attiva in ~88% dei "
    "casi controllati), ma l'effetto su E/F e' TRASCURABILE. Causa reale trovata (vedi note sotto): il "
    "floor di fine asta, non il ceiling di qualita', decide il bid nel 90% dei casi."
)

headers = ["Strategia", "Componenti Attivi", "FVM Medio", "FVM Mediana", "Deviazione Standard",
           "% Vittorie", "% Rose Complete", "Crediti Inutilizzati", "FVM per Credito", "Note"]
for i, h in enumerate(headers):
    sim.cell(row=4, column=1 + i, value=h)

componenti = {
    "A": "FVM puro",
    "B": "+ Performance storica",
    "C": "+ Scarsita' di ruolo",
    "D": "+ Quality Scarcity",
    "E": "+ Auction State (4 fasce, urgenza per fascia, fascia quality-based)",
    "F": "+ Competition Pressure (gated sulle riserve)",
}
note = {
    "A": "Invariata nel principio.",
    "B": "Effetto marginale confermato ancora una volta.",
    "C": "Ancora identica a B (la scarsita' di ruolo come definita raramente scende sotto 1).",
    "D": "Sostanzialmente invariata vs C.",
    "E": (
        "SOSTANZIALMENTE INVARIATA rispetto a v3 (807 -> 806 FVM medio). La correzione 5 e' implementata "
        "correttamente e si attiva regolarmente (su 200 aste-test: 4860 casi in cui un giocatore di fascia "
        "migliore capitava in uno slot peggiore, upgrade applicato in 4288 di questi = 88%), ma il suo "
        "effetto e' quasi sempre annullato da un meccanismo piu' forte: il floor di fine asta (budget "
        "residuo / slot residui * urgenza) decide il prezzo finale nel 89.6% delle decisioni di bid di "
        "questa strategia, contro solo il 10.4% in cui conta il ceiling 'di qualita'' (quello che la "
        "correzione 5 migliora). Il floor NON dipende affatto da quale giocatore e' chiamato: e' cieco "
        "alla qualita' tanto quanto lo era prima di ogni correzione."
    ),
    "F": "Sostanzialmente invariata (738 vs 744 in v3). Eredita lo stesso problema di E.",
}

for i, s in enumerate(("A", "B", "C", "D", "E", "F")):
    r = 5 + i
    d = R[s]
    sim.cell(row=r, column=1, value=f"Strategia {s}")
    sim.cell(row=r, column=2, value=componenti[s])
    sim.cell(row=r, column=3, value=round(d["fvm_medio"], 1))
    sim.cell(row=r, column=4, value=round(d["fvm_mediana"], 1))
    sim.cell(row=r, column=5, value=round(d["fvm_sd"], 1))
    sim.cell(row=r, column=6, value=round(d["win_pct"], 1))
    sim.cell(row=r, column=7, value=round(d["complete_pct"], 1))
    sim.cell(row=r, column=8, value=round(d["resid_medio"], 1))
    sim.cell(row=r, column=9, value=round(d["fvm_per_credito"], 3))
    sim.cell(row=r, column=10, value=note[s])

sim["A12"] = "STATO: correzione 5 (fascia quality-based) applicata e verificata attiva, ma NON risolutiva."
sim["A13"] = (
    "Diagnosi finale (misurata su 200 aste-test con instrumentazione dedicata, non solo ipotizzata): il "
    "89.6% delle decisioni di bid di Team1 in strategia E sono decise dal floor di fine asta, non dal "
    "ceiling che incorpora fasce/qualita'/correzioni 3-4-5. Il floor = (budget residuo squadra / slot "
    "residui totali) * fattore di ritmo * urgenza-fascia: non contiene alcun riferimento al FVM del "
    "giocatore chiamato in quel momento. Qualsiasi correzione al ceiling (comprese tutte quelle fatte "
    "finora: 3, 4, 4bis, 5) puo' quindi migliorare al massimo il 10% delle decisioni — il resto e' deciso "
    "da un meccanismo che tratta un giocatore fortissimo e uno scarso allo stesso modo, purche' capitino "
    "nello stesso slot/fascia/momento dell'asta."
)
sim["A14"] = (
    "Prossimo passo suggerito (non ancora implementato, da confermare): rendere anche il floor di fine "
    "asta sensibile al FVM del giocatore chiamato (es. scalarlo per lo stesso 'weight' di qualita' gia' "
    "calcolato per il ceiling, invece di applicarlo come valore fisso per slot). Senza questo, le "
    "correzioni al solo ceiling hanno un tetto strutturale di efficacia intorno al 10% delle decisioni."
)
sim["A15"] = "Nessuna modifica alle formule del motore live in questo passaggio."

wb.save(PATH)
print("Simulazioni v4 aggiornato.")
