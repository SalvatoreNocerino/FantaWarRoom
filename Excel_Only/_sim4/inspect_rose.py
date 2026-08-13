import openpyxl
wb = openpyxl.load_workbook(r"C:\Users\salva\Projects\FantaWarRoom\Excel_Only\Rose_simulazione_1.xlsx")
for sn in wb.sheetnames:
    ws = wb[sn]
    print("SHEET", sn, ws.dimensions)
    for r in range(1, 20):
        vals = [ws.cell(row=r, column=c).value for c in range(1, 6)]
        print(r, vals)
    print("freeze", ws.freeze_panes, "autofilter", ws.auto_filter.ref)
    print("---")
