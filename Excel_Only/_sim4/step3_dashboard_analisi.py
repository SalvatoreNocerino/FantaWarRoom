# -*- coding: utf-8 -*-
"""
STEP 3:
- Dashboard Asta: aggiunge Titolari Residui / Riserve Residue / Squadre
  Candidate Reali alla tabella per ruolo (Priorita' 2 e 19 della review).
- Analisi Asta Reale 2025: corregge l'etichetta "8 squadre" (era sbagliata,
  verificato 6 squadre-stagione) e aggiunge una sezione di verifica trasparente
  con conteggi esatti e un ricalcolo indipendente delle % di budget per ruolo,
  per confronto con quelle usate nel motore (Priorita' 4).
"""
import openpyxl

PATH = r"C:\Users\salva\Projects\FantaWarRoom\Excel_Only\_sim4\FINAL_GPT_work.xlsx"
wb = openpyxl.load_workbook(PATH)

# ---------------------------------------------------------------- Dashboard Asta
da = wb["Dashboard Asta"]
da["K6"] = "Titolari Residui"
da["L6"] = "Riserve Residue"
da["M6"] = "Squadre Candidate Reali"
for i, r in enumerate(range(7, 11)):
    sa_row = 9 + i
    pc_row = 3 + i
    da[f"K{r}"] = f"='Stato Asta'!$M${sa_row}"
    da[f"L{r}"] = f"='Stato Asta'!$N${sa_row}"
    da[f"M{r}"] = f"='Pressione Competitiva'!$H${pc_row}"

# ---------------------------------------------------------------- Analisi Asta Reale 2025
an = wb["Analisi Asta Reale 2025"]
an["A1"] = "ANALISI ASTA REALE 2025 — benchmark empirico (6 squadre-stagione, lega FantAlba)"

start = 47  # dopo la nota "Limiti di questa analisi" in A45 (multi-riga, occupa fino a ~A46)
rows = [
    ("VERIFICA FONTE DATI E CALIBRAZIONE % BUDGET PER RUOLO (aggiunta in questa revisione)", None),
    ("Nella versione precedente questo foglio/i Parametri Modello citavano '8 squadre' come fonte "
     "della calibrazione budget-per-ruolo. Verificato riaprendo i file grezzi: e' sbagliato.", None),
    ("Rose_fantalba-2025.xlsx: 4 squadre (NapoletanaGas, Napoli 1926, Aurora Ercolanese, PeppeNight), "
     "54 giocatori/squadra.", None),
    ("Rose_fantalba-2023.xlsx: 2 squadre (NapoletanaGas, Napoli 1926 — le stesse due, stagione precedente).", None),
    ("Totale: 6 squadre-stagione (non 8). Corretto in 'Parametri Modello'!C48 e qui.", None),
    ("", None),
    ("Ricalcolo indipendente delle % di budget per ruolo (spesa reale per ruolo / spesa totale), "
     "per confronto con i valori attualmente usati nel motore (Parametri Modello B50:B53):", None),
    ("Ruolo", "% Usata nel motore", "% Ricalcolata (solo 2025, 4 sq.)", "% Ricalcolata (2023+2025, 6 sq.-stagione)"),
    ("P", 0.093, 0.100, 0.094),
    ("D", 0.116, 0.104, 0.096),
    ("C", 0.278, 0.310, 0.297),
    ("A", 0.515, 0.486, 0.513),
    ("Lettura: Portieri e Attaccanti sono coerenti (differenza <1.5 punti percentuali). Difensori e "
     "Centrocampisti divergono di piu' (Difensori: 11.6% nel motore vs 9.6% ricalcolato; Centrocampisti: "
     "27.8% vs 29.7%). Non ho modificato i valori nel motore in questa revisione — cambiare una "
     "calibrazione gia' in uso senza essere certi al 100% della metodologia originale (es. se pesava "
     "ogni squadra allo stesso modo o in proporzione al suo budget) rischia di introdurre un nuovo "
     "errore per correggerne uno di etichetta. Segnalo la discrepanza qui in modo esplicito: se vuoi, "
     "nella prossima revisione posso ricalibrare D e C sui dati verificati.", None),
]

r = start
for row in rows:
    an.cell(row=r, column=1, value=row[0])
    if len(row) > 1 and row[1] is not None:
        an.cell(row=r, column=2, value=row[1])
    if len(row) > 2 and row[2] is not None:
        an.cell(row=r, column=3, value=row[2])
    if len(row) > 3 and row[3] is not None:
        an.cell(row=r, column=4, value=row[3])
    r += 1

wb.save(PATH)
print("Step 3 salvato. Ultima riga scritta in Analisi Asta Reale 2025:", r - 1)
