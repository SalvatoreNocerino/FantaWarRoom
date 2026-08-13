# -*- coding: utf-8 -*-
import json
import openpyxl

PATH = r"C:\Users\salva\Projects\FantaWarRoom\Excel_Only\Strategia_Asta_FINAL_GPT.xlsx"

with open(r"C:\Users\salva\Projects\FantaWarRoom\Excel_Only\_sim4\full_comparison_results2.json") as f:
    R = json.load(f)

wb = openpyxl.load_workbook(PATH)
sim = wb["Simulazioni"]

for rng in list(sim.merged_cells.ranges):
    if rng.min_row <= 20:
        sim.unmerge_cells(str(rng))
for row in sim.iter_rows(min_row=1, max_row=20):
    for cell in row:
        cell.value = None

sim["A1"] = "SIMULAZIONI — confronto strategie A-F v2 (correzioni applicate su richiesta utente)"
sim["A2"] = (
    "v2 rispetto alla versione precedente: (3) Scarsita' di mercato e Quality Scarcity premiano SOLO "
    "quando i giocatori scarseggiano, mai piu' sconto sotto il prezzo FVM base quando sono abbondanti; "
    "(4) Titolari/Riserve non e' piu' uno switch a 2 livelli ma a 4 fasce di qualita' decrescente (es. "
    "difensori: 3 titolari + 2 riserve alte + 2 riserve medie + 1 riserva bassa), ciascuna con budget "
    "proprio; (5) la pressione competitiva sulle riserve si applica solo se il giocatore chiamato ha un "
    "FVM nettamente piu' alto (soglia 1.3x) della media delle riserve rimaste nella sua fascia. "
    "Metodologia di confronto invariata: la mia squadra (Team1) usa la strategia in test, le altre 7 "
    "usano sempre il motore completo F. 1.000 aste per strategia."
)

headers = ["Strategia", "Componenti Attivi", "FVM Medio", "FVM Mediana", "Deviazione Standard",
           "% Vittorie", "% Rose Complete", "Crediti Inutilizzati", "FVM per Credito", "Note"]
for i, h in enumerate(headers):
    sim.cell(row=4, column=1 + i, value=h)

componenti = {
    "A": "FVM puro",
    "B": "+ Performance storica",
    "C": "+ Scarsita' di ruolo (corretta: solo premio, mai sconto)",
    "D": "+ Quality Scarcity (corretta: solo premio, mai sconto)",
    "E": "+ Auction State (4 fasce invece di 2)",
    "F": "+ Competition Pressure (gated sulle riserve)",
}
note = {
    "A": "Baseline invariata nel principio, valori leggermente diversi dalla v1 per via del nuovo seed e della struttura a 4 fasce disponibile anche quando non usata per il pricing (influenza il bookkeeping di mercato).",
    "B": "Ancora un effetto marginale (+0.9% su A), confermato: la Fantamedia aggiunge poco sopra un FVM che gia' la incorpora.",
    "C": (
        "IDENTICO a B in tutte le metriche. Causa verificata: il rapporto 'giocatori liberi di ruolo / "
        "slot residui lega' parte gia' molto sopra 1 (es. difensori 175 liberi per 64 slot = 2.73) e "
        "TENDE A SALIRE, non scendere, durante l'asta (il bacino di difensori mai comprati si svuota "
        "molto piu' lentamente degli slot). Con la correzione (3) che giustamente elimina lo sconto per "
        "abbondanza, questo segnale semplicemente non trova mai una vera scarsita' da premiare: non e' "
        "un problema della correzione, e' un limite di come avevo definito 'scarsita'' che la correzione "
        "ha fatto emergere. Andrebbe ricalibrato (es. confrontare con un sottoinsieme piu' realistico di "
        "giocatori 'ancora appetibili', non l'intero bacino)."
    ),
    "D": "Sostanzialmente invariata vs C, coerente con quanto gia' visto in v1.",
    "E": (
        "PEGGIORATA rispetto alla v1 (1056 -> 767 FVM medio). Le 4 fasce sono state costruite come "
        "richiesto (verificato: difensori 3+2+2+1) e la classificazione di un giocatore NON dipende dal "
        "prezzo. Ma ho trovato la causa reale del calo: il meccanismo 'non sprecare crediti' di fine "
        "asta forza offerte alte in base al rapporto budget-rimasto/slot-rimasti della squadra, SENZA "
        "guardare se lo slot che si sta riempiendo e' di qualita' alta o bassa. Esempio verificato: un "
        "giocatore da 4 di FVM (Anjorin) comprato a 18 crediti nella fascia 'titolari' perche' la "
        "squadra doveva ancora spendere quel budget, non perche' valesse 18 crediti. Questo meccanismo "
        "esisteva gia' prima delle 3 correzioni richieste: non l'ho toccato perche' non era tra le 3 "
        "correzioni discusse, ma e' la causa principale per cui titolari/riserve a 4 fasce non basta da "
        "sola a risolvere il problema."
    ),
    "F": (
        "Ancora la peggiore (751 FVM medio, 12.8% vittorie, sostanzialmente il livello di una squadra "
        "media senza vantaggio, 1/8=12.5%). La correzione (5) riduce le aste inutili sulle riserve "
        "intercambiabili ma non basta a compensare il problema di fondo ereditato da E (vedi nota sopra)."
    ),
}

for i, s in enumerate(("A", "B", "C", "D", "E", "F")):
    r = 5 + i
    d = R[s]
    sim.cell(row=r, column=1, value=f"Strategia {s}")
    sim.cell(row=r, column=2, value=componenti[s])
    sim.cell(row=r, column=3, value=round(d["fvm_medio"], 1))
    sim.cell(row=r, column=4, value=round(d["fvm_mediana"], 1))
    sim.cell(row=r, column=5, value=round(d["fvm_sd"], 1))
    sim.cell(row=r, column=6, value=round(d["win_pct"], 1))
    sim.cell(row=r, column=7, value=round(d["complete_pct"], 1))
    sim.cell(row=r, column=8, value=round(d["resid_medio"], 1))
    sim.cell(row=r, column=9, value=round(d["fvm_per_credito"], 3))
    sim.cell(row=r, column=10, value=note[s])

sim["A12"] = "STATO DELLE 3 CORREZIONI RICHIESTE"
sim["A13"] = "(3) Fatta e verificata: nessuno sconto sotto il FVM base per abbondanza, solo premio per scarsita'."
sim["A14"] = "(4) Fatta e verificata: 4 fasce a qualita' decrescente invece di 2 (es. difensori 3+2+2+1 come da esempio)."
sim["A15"] = "(5) Fatta e verificata: pressione competitiva sulle riserve solo se il giocatore e' >=1.3x la media della sua fascia."
sim["A16"] = (
    "PROBLEMA RESIDUO (non richiesto in questa correzione, segnalato per una prossima): il meccanismo "
    "'non sprecare crediti' di fine asta ignora la qualita' dello slot che sta riempiendo. E' la causa "
    "principale per cui Auction State (E) e Competition Pressure (F) restano sotto le strategie piu' "
    "semplici anche dopo le 3 correzioni. Andrebbe reso sensibile alla fascia (es. urgenza forte solo "
    "sui titolari mancanti, urgenza debole/nulla su riserve gia' numerose)."
)

wb.save(PATH)
print("Simulazioni v2 aggiornato.")
