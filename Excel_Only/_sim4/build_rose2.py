# -*- coding: utf-8 -*-
"""
Rigenera Rose_simulazione_1.xlsx con il motore corretto (v3 + "brucia budget"
di fine asta). Due fogli:
  - "Strategia A" = motore semplice, FVM puro per ruolo, nessuna suddivisione
    titolari/riserve, nessuna scarsita' dinamica (baseline).
  - "Strategia F" = motore completo (sim3): titolari/riserve pianificati per
    ruolo + scarsita' dinamica. Stesso meccanismo "brucia budget" di fine asta
    applicato a entrambi, perche' rappresenta la dinamica competitiva generale
    (i crediti non spesi sono persi), non una differenza tra i due motori.
Stesso seed (42) per entrambi i fogli, per confrontare i due motori sulla
STESSA asta (stessi giocatori disponibili, stesso ordine di chiamata).
"""
import random
import openpyxl
from openpyxl.utils import get_column_letter

import sys
sys.path.insert(0, r"C:\Users\salva\Projects\FantaWarRoom\Excel_Only\_sim4")
from sim3 import PLAYERS, TIERS, ROSTER, N_TEAMS, BUDGET, ROLE_BUDGET_SHARE, run_auction, summarize

SEED = 42

# ---------------------------------------------------------------- Strategia A
def run_auction_simple(players, seed, noise_sigma=0.15, noise_clip=(0.7, 1.4)):
    rng = random.Random(seed)
    teams = {f"Team{i+1}": {"budget": BUDGET, "roster": [], "bought": {r: 0 for r in ROSTER}}
             for i in range(N_TEAMS)}
    remaining_fvm = {r: sum(p["fvm"] for p in players if p["ruolo"] == r) for r in ROSTER}
    remaining_n = {r: sum(1 for p in players if p["ruolo"] == r) for r in ROSTER}
    order = players[:]
    rng.shuffle(order)
    league_slots_left = {r: N_TEAMS * ROSTER[r] for r in ROSTER}
    TOTAL_SLOTS_TEAM = sum(ROSTER.values())
    TOTAL_SLOTS_LEAGUE = TOTAL_SLOTS_TEAM * N_TEAMS
    slots_filled_league = 0

    for p in order:
        role = p["ruolo"]
        if league_slots_left[role] <= 0:
            continue
        pace_league = slots_filled_league / TOTAL_SLOTS_LEAGUE if TOTAL_SLOTS_LEAGUE else 0
        bids = {}
        for tname, t in teams.items():
            slots_role_left = ROSTER[role] - t["bought"][role]
            if slots_role_left <= 0:
                continue
            slots_total_left = sum(ROSTER[r] - t["bought"][r] for r in ROSTER)
            other_open = slots_total_left - slots_role_left
            reserve = other_open + max(0, slots_role_left - 1)
            spendable = max(0, t["budget"] - reserve)
            if spendable <= 0:
                continue
            role_budget = ROLE_BUDGET_SHARE[role] * BUDGET
            spent_role = sum(x["prezzo"] for x in t["roster"] if x["ruolo"] == role)
            remaining_target = max(1, role_budget - spent_role)
            per_slot = remaining_target / slots_role_left
            avg_fvm = (remaining_fvm[role] / remaining_n[role]) if remaining_n[role] > 0 else 1
            weight = p["fvm"] / max(1, avg_fvm) if avg_fvm else 1.0
            weight = max(0.15, min(4.0, weight))
            ceiling_planned = per_slot * weight

            pace_factor = 0.35 + 1.15 * pace_league
            endgame_floor = (t["budget"] / max(1, slots_total_left)) * pace_factor
            ceiling_final = max(ceiling_planned, endgame_floor)

            noise = min(noise_clip[1], max(noise_clip[0], rng.lognormvariate(0, noise_sigma)))
            bid = min(ceiling_final * noise, spendable)
            if bid >= 1:
                bids[tname] = bid
        if not bids:
            continue
        winner = max(bids, key=bids.get)
        sorted_bids = sorted(bids.values(), reverse=True)
        price = int(round(sorted_bids[1])) if len(sorted_bids) > 1 else 1
        price = max(1, min(price, teams[winner]["budget"]))
        wt = teams[winner]
        wt["budget"] -= price
        wt["bought"][role] += 1
        wt["roster"].append({**p, "prezzo": price})
        league_slots_left[role] -= 1
        slots_filled_league += 1
        remaining_fvm[role] -= p["fvm"]
        remaining_n[role] -= 1
    return teams


def summarize_simple(teams):
    rows = []
    for tname, t in teams.items():
        fvm_tot = sum(p["fvm"] for p in t["roster"])
        spent = BUDGET - t["budget"]
        rows.append({"team": tname, "fvm_totale": round(fvm_tot, 1), "budget_speso": spent,
                     "budget_residuo": t["budget"], "n_giocatori": len(t["roster"])})
    return rows


print("Eseguo Strategia A (FVM puro, seed=%d)..." % SEED)
teams_a = run_auction_simple(PLAYERS, SEED)
rows_a = summarize_simple(teams_a)
spend_a = sum(r["budget_speso"] for r in rows_a) / (N_TEAMS * BUDGET) * 100
print("  spesa media: %.1f%%  FVM medio: %.1f" % (spend_a, sum(r["fvm_totale"] for r in rows_a) / N_TEAMS))

print("Eseguo Strategia F (motore completo, seed=%d)..." % SEED)
teams_f = run_auction(PLAYERS, TIERS, SEED)
rows_f = summarize(teams_f)
spend_f = sum(r["budget_speso"] for r in rows_f) / (N_TEAMS * BUDGET) * 100
print("  spesa media: %.1f%%  FVM medio: %.1f" % (spend_f, sum(r["fvm_totale"] for r in rows_f) / N_TEAMS))

ROLE_ORDER = {"P": 0, "D": 1, "C": 2, "A": 3}

# ---------------------------------------------------------------- Scrivi file
SRC = r"C:\Users\salva\Projects\FantaWarRoom\Excel_Only\Rose_simulazione_1.xlsx"
OUT = r"C:\Users\salva\Projects\FantaWarRoom\Excel_Only\_sim4\Rose_simulazione_1_FIX.xlsx"
wb = openpyxl.load_workbook(SRC)

def fill_sheet(ws, teams, note2):
    # righe 1-17 restano invariate (titoli, intestazioni, formule di riepilogo)
    ws.cell(row=2, column=1).value = note2
    r = 18
    for tname in sorted(teams.keys(), key=lambda x: int(x.replace("Team", ""))):
        t = teams[tname]
        roster_sorted = sorted(t["roster"], key=lambda p: (ROLE_ORDER.get(p["ruolo"], 9), -p["prezzo"]))
        for p in roster_sorted:
            ws.cell(row=r, column=1).value = tname
            ws.cell(row=r, column=2).value = p["ruolo"]
            ws.cell(row=r, column=3).value = p["nome"]
            ws.cell(row=r, column=4).value = p["prezzo"]
            ws.cell(row=r, column=5).value = p["fvm"]
            r += 1
    return r - 1

ws_a = wb["Strategia A"]
last_a = fill_sheet(
    ws_a, teams_a,
    "Strategia A = FVM puro (allocazione proporzionale al FVM per ruolo, nessuna suddivisione titolari/riserve, "
    "nessuna scarsita' dinamica). Include il meccanismo 'brucia budget' di fine asta (i crediti non spesi sono "
    "persi, quindi la spinta a spendere cresce col progredire dell'asta). Stessa asta (seed 42), 8 squadre, "
    "budget 1000, rosa 3P-8D-8C-6A."
)
ws_f = wb["Strategia F"]
last_f = fill_sheet(
    ws_f, teams_f,
    "Strategia F = motore completo usato in Strategia_Asta_FINAL_Cowork.xlsx: titolari/riserve pianificati per "
    "ruolo + scarsita' dinamica + meccanismo 'brucia budget' di fine asta. Stessa asta (seed 42), 8 squadre, "
    "budget 1000, rosa 3P-8D-8C-6A."
)
print("Ultima riga scritta A:", last_a, " F:", last_f)

wb.save(OUT)
print("Salvato:", OUT)
