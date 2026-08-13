import openpyxl, io
wb = openpyxl.load_workbook(r"C:\Users\salva\Projects\FantaWarRoom\Excel_Only\Dati\Rose_fantalba-2025.xlsx", data_only=True)
ws = wb["TutteLeRose"]
out = io.open(r"C:\Users\salva\Projects\FantaWarRoom\Excel_Only\_sim4\dump_2025_full.txt", "w", encoding="utf-8")
out.write("dim=%s max_col=%s max_row=%s\n" % (ws.dimensions, ws.max_column, ws.max_row))
for r in range(1, 10):
    vals = [ws.cell(row=r, column=c).value for c in range(1, 9)]
    out.write("row%d: %s\n" % (r, vals))
out.write("... ultima riga:\n")
for r in range(ws.max_row - 3, ws.max_row + 1):
    vals = [ws.cell(row=r, column=c).value for c in range(1, 9)]
    out.write("row%d: %s\n" % (r, vals))
out.close()
print("done")
