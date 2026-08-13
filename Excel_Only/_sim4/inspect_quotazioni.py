import openpyxl, io
wb = openpyxl.load_workbook(r"C:\Users\salva\Projects\FantaWarRoom\Excel_Only\Dati\Quotazioni_Fantacalcio_Stagione_2025_26 (2).xlsx", data_only=True)
out = io.open(r"C:\Users\salva\Projects\FantaWarRoom\Excel_Only\_sim4\dump_quotazioni.txt", "w", encoding="utf-8")
out.write("sheets: %s\n" % wb.sheetnames)
for sn in wb.sheetnames:
    ws = wb[sn]
    out.write("sheet %s dim=%s\n" % (sn, ws.dimensions))
    for r in range(1, 5):
        vals = [ws.cell(row=r, column=c).value for c in range(1, min(ws.max_column, 20) + 1)]
        out.write("  row%d: %s\n" % (r, vals))
out.close()
print("done")
