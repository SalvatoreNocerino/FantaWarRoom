import openpyxl, io
wb = openpyxl.load_workbook(r"C:\Users\salva\Projects\FantaWarRoom\Excel_Only\Strategia_Asta_FINAL_GPT.xlsx", data_only=True)
ws = wb["Input Statistiche 2025_26"]
out = io.open(r"C:\Users\salva\Projects\FantaWarRoom\Excel_Only\_sim4\dump_statistiche.txt", "w", encoding="utf-8")
out.write("dim=%s\n" % ws.dimensions)
for r in range(1, 5):
    vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
    out.write("row%d: %s\n" % (r, vals))
out.close()
print("done")
