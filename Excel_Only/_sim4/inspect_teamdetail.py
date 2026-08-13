import openpyxl, io
wb = openpyxl.load_workbook(r"C:\Users\salva\Projects\FantaWarRoom\Excel_Only\_sim4\FINAL_GPT_work.xlsx")
ws = wb["TeamDetail"]
out = io.open(r"C:\Users\salva\Projects\FantaWarRoom\Excel_Only\_sim4\dump_teamdetail.txt", "w", encoding="utf-8")
out.write("DIM %s\n" % ws.dimensions)
for r in range(1, min(ws.max_row, 20) + 1):
    for c in range(1, ws.max_column + 1):
        cell = ws.cell(row=r, column=c)
        v = cell.value
        if v is not None:
            if hasattr(v, "text"):
                v = "ARRAY:" + str(v.text)
            out.write("%s %s\n" % (cell.coordinate, repr(v)[:250]))
out.close()
print("done")
