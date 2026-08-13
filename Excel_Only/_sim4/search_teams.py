import openpyxl, io
wb = openpyxl.load_workbook(r"C:\Users\salva\Projects\FantaWarRoom\Excel_Only\_sim4\FINAL_GPT_work.xlsx")
out = io.open(r"C:\Users\salva\Projects\FantaWarRoom\Excel_Only\_sim4\dump_teamcount_mentions.txt", "w", encoding="utf-8")
for sn in wb.sheetnames:
    ws = wb[sn]
    for row in ws.iter_rows():
        for c in row:
            v = c.value
            if hasattr(v, "text"):
                v = v.text
            if isinstance(v, str) and ("squadr" in v.lower() or "fantalba" in v.lower()):
                out.write(f"{sn} {c.coordinate}: {v[:300]}\n")
out.close()
print("done")
