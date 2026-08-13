import openpyxl, io
out = io.open(r"C:\Users\salva\Projects\FantaWarRoom\Excel_Only\_sim4\dump_real_data_check.txt", "w", encoding="utf-8")

for fn in [r"C:\Users\salva\Projects\FantaWarRoom\Excel_Only\Dati\Rose_fantalba-2025.xlsx",
           r"C:\Users\salva\Projects\FantaWarRoom\Excel_Only\Dati\Rose_fantalba-2023.xlsx"]:
    wb = openpyxl.load_workbook(fn, data_only=True)
    out.write(f"\n=== {fn} ===\n")
    for sn in wb.sheetnames:
        ws = wb[sn]
        out.write(f"  sheet {sn} dim={ws.dimensions}\n")
    ws = wb[wb.sheetnames[0]]
    # stampa prima riga e prime colonne per capire la struttura "a blocchi"
    for r in range(1, min(ws.max_row, 5) + 1):
        vals = [ws.cell(row=r, column=c).value for c in range(1, min(ws.max_column, 30) + 1)]
        out.write(f"  row{r}: {vals}\n")
out.close()
print("done")
