# -*- coding: utf-8 -*-
"""
Simulatore v3: ogni squadra pianifica PRIMA dell'asta quanto vuole spendere
per ruolo (target %) e quanto di quel budget va ai titolari vs alle riserve.
Durante l'asta ricalcola, ad ogni chiamata, il proprio "prezzo giusto" per
QUEL giocatore = (budget di ruolo/tier ancora disponibile) / (slot di
ruolo/tier ancora da riempire) * (forza relativa del giocatore nel tier).
Non c'e' piu' un valore FVM globale fisso uguale per tutti: ogni squadra
rivaluta in base a quanto le e' rimasto e quanto le serve ancora.
"""
import random, json
import openpyxl

SRC = r"C:\Users\salva\Projects\FantaWarRoom\Excel_Only\Strategia_Asta_FINAL_Cowork.xlsx"
N_TEAMS = 8
BUDGET = 1000
ROSTER = {"P": 3, "D": 8, "C": 8, "A": 6}
TITOLARI = {"P": 1, "D": 3, "C": 4, "A": 3}
ROLE_BUDGET_SHARE = {"P": 0.093, "D": 0.116, "C": 0.278, "A": 0.515}
TITOLARI_BUDGET_SHARE = {"P": 0.88, "D": 0.85, "C": 0.95, "A": 0.93}

def load_players():
    wb = openpyxl.load_workbook(SRC, data_only=True)
    ws = wb["Input Giocatori 2026_27"]
    players = []
    for r in range(3, ws.max_row + 1):
        pid = ws.cell(row=r, column=1).value
        ruolo = ws.cell(row=r, column=2).value
        nome = ws.cell(row=r, column=3).value
        fvm = ws.cell(row=r, column=7).value
        if pid and ruolo in ROSTER and nome:
            players.append({"id": pid, "ruolo": ruolo, "nome": nome, "fvm": fvm or 0})
    return players

PLAYERS = load_players()
print("Giocatori:", len(PLAYERS))

def build_tiers(players):
    tiers = {}
    for role in ROSTER:
        role_players = sorted([p for p in players if p["ruolo"] == role], key=lambda p: -p["fvm"])
        pool_n = N_TEAMS * ROSTER[role]
        tit_n = N_TEAMS * TITOLARI[role]
        titolari_ids = {p["id"] for p in role_players[:tit_n]}
        riserve_ids = {p["id"] for p in role_players[tit_n:pool_n]}
        tiers[role] = {
            "titolari_ids": titolari_ids, "riserve_ids": riserve_ids,
            "titolari_fvm_avg": (sum(p["fvm"] for p in role_players[:tit_n]) / max(1, tit_n)),
            "riserve_fvm_avg": (sum(p["fvm"] for p in role_players[tit_n:pool_n]) / max(1, pool_n - tit_n)) or 1,
        }
    return tiers

TIERS = build_tiers(PLAYERS)
for r, t in TIERS.items():
    print(r, "titolari:", len(t["titolari_ids"]), "avg fvm:", round(t["titolari_fvm_avg"]),
          " riserve:", len(t["riserve_ids"]), "avg fvm:", round(t["riserve_fvm_avg"]))

def run_auction(players, tiers, seed, noise_sigma=0.15, noise_clip=(0.7, 1.4)):
    rng = random.Random(seed)
    teams = {}
    for i in range(N_TEAMS):
        teams[f"Team{i+1}"] = {
            "budget": BUDGET, "roster": [],
            "tit_bought": {r: 0 for r in ROSTER}, "ris_bought": {r: 0 for r in ROSTER},
            "spent_tit": {r: 0 for r in ROSTER}, "spent_ris": {r: 0 for r in ROSTER},
        }
    remaining_fvm = {r: {"titolari": sum(p["fvm"] for p in players if p["id"] in tiers[r]["titolari_ids"]),
                          "riserve": sum(p["fvm"] for p in players if p["id"] in tiers[r]["riserve_ids"])}
                      for r in ROSTER}
    remaining_n = {r: {"titolari": len(tiers[r]["titolari_ids"]), "riserve": len(tiers[r]["riserve_ids"])}
                   for r in ROSTER}

    order = players[:]
    rng.shuffle(order)
    league_slots_left = {r: N_TEAMS * ROSTER[r] for r in ROSTER}
    TOTAL_SLOTS_TEAM = sum(ROSTER.values())
    TOTAL_SLOTS_LEAGUE = TOTAL_SLOTS_TEAM * N_TEAMS
    slots_filled_league = 0

    def player_tier(role, pid):
        if pid in tiers[role]["titolari_ids"]:
            return "titolari"
        if pid in tiers[role]["riserve_ids"]:
            return "riserve"
        return "floor"

    def bid_ceiling(team, role, pid, fvm, market_tier):
        t = team
        needs_tit = t["tit_bought"][role] < TITOLARI[role]
        use_tier = "titolari" if (needs_tit and market_tier != "floor") else "riserve"
        # una squadra che ha gia' i suoi titolari non paga piu' prezzo da titolare,
        # anche se il giocatore e' oggettivamente di fascia alta sul mercato
        role_budget = ROLE_BUDGET_SHARE[role] * BUDGET
        if use_tier == "titolari":
            target = role_budget * TITOLARI_BUDGET_SHARE[role]
            spent = t["spent_tit"][role]
            slots_needed = TITOLARI[role] - t["tit_bought"][role]
            avg_fvm = (remaining_fvm[role]["titolari"] / remaining_n[role]["titolari"]) \
                if remaining_n[role]["titolari"] > 0 else tiers[role]["titolari_fvm_avg"]
        else:
            target = role_budget * (1 - TITOLARI_BUDGET_SHARE[role])
            spent = t["spent_ris"][role]
            slots_needed = (ROSTER[role] - TITOLARI[role]) - t["ris_bought"][role]
            if market_tier == "floor":
                avg_fvm = max(1, tiers[role]["riserve_fvm_avg"] * 0.3)
            else:
                avg_fvm = (remaining_fvm[role]["riserve"] / remaining_n[role]["riserve"]) \
                    if remaining_n[role]["riserve"] > 0 else tiers[role]["riserve_fvm_avg"]
        if slots_needed <= 0:
            return None, use_tier
        remaining_target = max(1, target - spent)
        per_slot = remaining_target / slots_needed
        weight = fvm / max(1, avg_fvm) if avg_fvm else 1.0
        weight = max(0.15, min(4.0, weight))
        return per_slot * weight, use_tier

    for p in order:
        role = p["ruolo"]
        if league_slots_left[role] <= 0:
            continue
        market_tier = player_tier(role, p["id"])

        # scarsita' dinamica: quante squadre cercano ancora un titolare/riserva in
        # questo ruolo rispetto a quanti giocatori di quel tier restano sul mercato
        n_need_tit = sum(1 for t in teams.values() if t["tit_bought"][role] < TITOLARI[role])
        n_need_ris = sum(1 for t in teams.values()
                          if (ROSTER[role] - TITOLARI[role]) - t["ris_bought"][role] > 0)
        scarcity_tit = min(2.5, max(0.6, n_need_tit / max(1, remaining_n[role]["titolari"])))
        scarcity_ris = min(2.5, max(0.6, n_need_ris / max(1, remaining_n[role]["riserve"] + 1)))

        bids, tiers_used = {}, {}
        pace_league = slots_filled_league / TOTAL_SLOTS_LEAGUE if TOTAL_SLOTS_LEAGUE else 0
        for tname, t in teams.items():
            slots_total_left = (TITOLARI[role] - t["tit_bought"][role]) + \
                                ((ROSTER[role] - TITOLARI[role]) - t["ris_bought"][role])
            if slots_total_left <= 0:
                continue
            other_open = sum((TITOLARI[r] - t["tit_bought"][r]) + ((ROSTER[r] - TITOLARI[r]) - t["ris_bought"][r])
                              for r in ROSTER if r != role)
            reserve = other_open + max(0, slots_total_left - 1)
            spendable = max(0, t["budget"] - reserve)
            if spendable <= 0:
                continue
            ceiling, use_tier = bid_ceiling(t, role, p["id"], p["fvm"], market_tier)
            if ceiling is None:
                continue
            scarcity_mult = scarcity_tit if use_tier == "titolari" else scarcity_ris
            ceiling_planned = ceiling * scarcity_mult

            # "brucia budget": i crediti non spesi a fine asta sono persi, quindi
            # una squadra con ancora molto budget e pochi slot totali rimasti (di
            # QUALSIASI ruolo) accetta di pagare anche una riserva piu' di quanto
            # pianificato, soprattutto verso la fine dell'asta. Il pavimento
            # cresce col progredire dell'asta (pace_league 0->1) cosi' da non
            # distorcere la scoperta del prezzo nella prima parte dell'asta.
            own_slots_left_all = slots_total_left + other_open
            pace_factor = 0.35 + 1.15 * pace_league
            endgame_floor = (t["budget"] / max(1, own_slots_left_all)) * pace_factor
            ceiling_final = max(ceiling_planned, endgame_floor)

            noise = min(noise_clip[1], max(noise_clip[0], rng.lognormvariate(0, noise_sigma)))
            bid = min(ceiling_final * noise, spendable)
            if bid >= 1:
                bids[tname] = bid
                tiers_used[tname] = use_tier
        if not bids:
            continue
        winner = max(bids, key=bids.get)
        sorted_bids = sorted(bids.values(), reverse=True)
        price = int(round(sorted_bids[1])) if len(sorted_bids) > 1 else 1
        price = max(1, min(price, teams[winner]["budget"]))
        wt = teams[winner]
        use_tier = tiers_used[winner]
        wt["budget"] -= price
        if use_tier == "titolari":
            wt["tit_bought"][role] += 1
            wt["spent_tit"][role] += price
        else:
            wt["ris_bought"][role] += 1
            wt["spent_ris"][role] += price
        wt["roster"].append({**p, "prezzo": price, "tier": use_tier})
        league_slots_left[role] -= 1
        slots_filled_league += 1
        if market_tier != "floor":
            remaining_fvm[role][market_tier] -= p["fvm"]
            remaining_n[role][market_tier] -= 1
    return teams
print("Motore v3 pronto.")

def summarize(teams):
    rows = []
    for tname, t in teams.items():
        fvm_tot = sum(p["fvm"] for p in t["roster"])
        spent = BUDGET - t["budget"]
        rows.append({"team": tname, "fvm_totale": round(fvm_tot, 1), "budget_speso": spent,
                     "budget_residuo": t["budget"], "n_giocatori": len(t["roster"])})
    return rows
