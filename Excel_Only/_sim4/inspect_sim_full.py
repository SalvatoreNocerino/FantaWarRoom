import openpyxl, io
wb = openpyxl.load_workbook(r"C:\Users\salva\Projects\FantaWarRoom\Excel_Only\Strategia_Asta_FINAL_GPT.xlsx")
ws = wb["Simulazioni"]
out = io.open(r"C:\Users\salva\Projects\FantaWarRoom\Excel_Only\_sim4\dump_sim_full.txt", "w", encoding="utf-8")
out.write("dim=%s merged=%s\n" % (ws.dimensions, list(ws.merged_cells.ranges)))
for r in range(1, ws.max_row + 1):
    v = ws.cell(row=r, column=1).value
    if v is not None:
        out.write("A%d: %s\n" % (r, str(v)[:200]))
out.close()
print("done")
