import openpyxl, io
wb = openpyxl.load_workbook(r"C:\Users\salva\Projects\FantaWarRoom\Excel_Only\_sim4\FINAL_GPT_work.xlsx")

def dump(sn, path, maxr=None, maxc=None):
    ws = wb[sn]
    out = io.open(path, "w", encoding="utf-8")
    out.write("DIM %s\n" % ws.dimensions)
    mr = maxr or ws.max_row
    mc = maxc or ws.max_column
    for r in range(1, mr + 1):
        for c in range(1, mc + 1):
            cell = ws.cell(row=r, column=c)
            v = cell.value
            if v is not None:
                if hasattr(v, "text"):
                    v = "ARRAY:" + str(v.text)
                out.write("%s %s\n" % (cell.coordinate, repr(v)[:250]))
    out.close()

dump("La Mia Rosa", r"C:\Users\salva\Projects\FantaWarRoom\Excel_Only\_sim4\dump_mia_rosa.txt", maxr=30)
dump("Pressione Competitiva", r"C:\Users\salva\Projects\FantaWarRoom\Excel_Only\_sim4\dump_pressione.txt")
dump("Dashboard Asta", r"C:\Users\salva\Projects\FantaWarRoom\Excel_Only\_sim4\dump_dashboard.txt")
dump("Giocatori Da Chiamare", r"C:\Users\salva\Projects\FantaWarRoom\Excel_Only\_sim4\dump_dachiamare.txt", maxr=15)
dump("Analisi Asta Reale 2025", r"C:\Users\salva\Projects\FantaWarRoom\Excel_Only\_sim4\dump_analisi.txt")
print("done")
