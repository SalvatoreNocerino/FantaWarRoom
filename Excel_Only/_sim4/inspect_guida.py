import openpyxl, io
wb = openpyxl.load_workbook(r"C:\Users\salva\Projects\FantaWarRoom\Excel_Only\_sim4\FINAL_GPT_work.xlsx")
ws = wb["Guida"]
out = io.open(r"C:\Users\salva\Projects\FantaWarRoom\Excel_Only\_sim4\dump_guida.txt", "w", encoding="utf-8")
out.write("DIM %s max_row=%s\n" % (ws.dimensions, ws.max_row))
for r in range(1, ws.max_row + 1):
    v = ws.cell(row=r, column=1).value
    if v is not None:
        out.write("A%d: %s\n" % (r, v))
out.close()
print("done")
