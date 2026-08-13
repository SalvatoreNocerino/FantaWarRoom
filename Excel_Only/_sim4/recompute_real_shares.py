import openpyxl, io

out = io.open(r"C:\Users\salva\Projects\FantaWarRoom\Excel_Only\_sim4\dump_recompute_shares.txt", "w", encoding="utf-8")

def parse_2025(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["TutteLeRose"]
    teams = {}
    # blocchi di 6 colonne: Ruolo,Calciatore,Squadra,Costo,vuoto,vuoto  (nome squadra in riga1)
    c = 1
    while c <= ws.max_column:
        name = ws.cell(row=1, column=c).value
        if name and isinstance(name, str) and name.strip():
            role_c, cost_c = c, c + 3
            rows = []
            r = 3
            while True:
                ruolo = ws.cell(row=r, column=role_c).value
                costo = ws.cell(row=r, column=cost_c).value
                if ruolo is None and costo is None:
                    # tollera un paio di righe vuote poi stop
                    if r > ws.max_row or ws.cell(row=r + 1, column=role_c).value is None:
                        break
                if ruolo:
                    try:
                        costo_v = float(costo) if costo not in (None, "") else None
                    except Exception:
                        costo_v = None
                    rows.append((ruolo, costo_v))
                r += 1
                if r > ws.max_row:
                    break
            teams[name] = rows
        c += 6
    return teams

def parse_2023(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["TutteLeRose"]
    teams = {}
    c = 1
    while c <= ws.max_column:
        name = ws.cell(row=5, column=c).value
        if name and isinstance(name, str) and name.strip():
            role_c, cost_c = c, c + 3
            rows = []
            r = 7
            while r <= ws.max_row:
                ruolo = ws.cell(row=r, column=role_c).value
                costo = ws.cell(row=r, column=cost_c).value
                if ruolo is None:
                    break
                try:
                    costo_v = float(costo) if costo not in (None, "") else None
                except Exception:
                    costo_v = None
                rows.append((ruolo, costo_v))
                r += 1
            teams[name] = rows
        c += 5
    return teams

t2025 = parse_2025(r"C:\Users\salva\Projects\FantaWarRoom\Excel_Only\Dati\Rose_fantalba-2025.xlsx")
t2023 = parse_2023(r"C:\Users\salva\Projects\FantaWarRoom\Excel_Only\Dati\Rose_fantalba-2023.xlsx")

out.write("Squadre 2025: %s\n" % list(t2025.keys()))
for k, v in t2025.items():
    out.write("  %s: %d giocatori, costo totale=%s\n" % (k, len(v), sum(x[1] or 0 for x in v)))
out.write("Squadre 2023: %s\n" % list(t2023.keys()))
for k, v in t2023.items():
    out.write("  %s: %d giocatori, costo totale=%s\n" % (k, len(v), sum(x[1] or 0 for x in v)))

def role_share(all_teams):
    role_sum = {"P": 0, "D": 0, "C": 0, "A": 0}
    total = 0
    n_teams = 0
    for tname, rows in all_teams.items():
        for ruolo, costo in rows:
            if ruolo in role_sum and costo is not None:
                role_sum[ruolo] += costo
                total += costo
        n_teams += 1
    out.write("  n_squadre=%d totale_speso=%.0f\n" % (n_teams, total))
    for r in role_sum:
        out.write("  %s: speso=%.0f share=%.4f\n" % (r, role_sum[r], role_sum[r] / total if total else 0))
    return role_sum, total, n_teams

out.write("\n--- Solo 2025 (4 squadre) ---\n")
role_share(t2025)

out.write("\n--- Solo 2023 (%d squadre) ---\n" % len(t2023))
role_share(t2023)

out.write("\n--- Combinato 2023+2025 (%d squadre-stagione) ---\n" % (len(t2025) + len(t2023)))
combined = {**{"2025_" + k: v for k, v in t2025.items()}, **{"2023_" + k: v for k, v in t2023.items()}}
role_share(combined)

out.close()
print("done")
