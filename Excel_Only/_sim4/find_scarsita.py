import openpyxl
wb = openpyxl.load_workbook(r"C:\Users\salva\Projects\FantaWarRoom\Excel_Only\_sim4\FINAL_GPT_work.xlsx")
for sn in wb.sheetnames:
    ws = wb[sn]
    for row in ws.iter_rows():
        for c in row:
            if isinstance(c.value, str) and ("Scarsit" in c.value or "Quality" in c.value or "Titolari Residui" in c.value or "Riserve Residue" in c.value):
                print(sn, c.coordinate, repr(c.value)[:120])
