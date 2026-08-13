# -*- coding: utf-8 -*-
"""
Ricalcolo completo dell'analisi asta reale con i file ORIGINALI corretti
(caricati dall'utente): Rose_fantalba-2025 ha 8 squadre (non 4 come nel file
che era in Dati/), Rose_fantalba-2023 ne ha 2 (confermato, era gia' corretto).
"""
import openpyxl, io, statistics

DATI = r"C:\Users\salva\Projects\FantaWarRoom\Excel_Only\Dati"

def parse_2025(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["TutteLeRose"]
    teams = {}
    c = 1
    while c <= ws.max_column:
        name = ws.cell(row=2, column=c).value
        if name and isinstance(name, str) and name.strip():
            role_c, player_c, cost_c = c, c + 1, c + 3
            rows = []
            r = 4
            while r <= ws.max_row:
                ruolo = ws.cell(row=r, column=role_c).value
                player = ws.cell(row=r, column=player_c).value
                costo = ws.cell(row=r, column=cost_c).value
                if ruolo and player:
                    try:
                        costo_v = float(costo) if costo not in (None, "") else None
                    except Exception:
                        costo_v = None
                    rows.append((ruolo, str(player).strip(), costo_v))
                r += 1
            teams[name] = rows
        c += 4
    return teams

def parse_2023(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["TutteLeRose"]
    teams = {}
    c = 1
    while c <= ws.max_column:
        name = ws.cell(row=5, column=c).value
        if name and isinstance(name, str) and name.strip():
            role_c, player_c, cost_c = c, c + 1, c + 3
            rows = []
            r = 7
            while r <= ws.max_row:
                ruolo = ws.cell(row=r, column=role_c).value
                player = ws.cell(row=r, column=player_c).value
                costo = ws.cell(row=r, column=cost_c).value
                if ruolo is None or player is None:
                    break
                try:
                    costo_v = float(costo) if costo not in (None, "") else None
                except Exception:
                    costo_v = None
                rows.append((ruolo, str(player).strip(), costo_v))
                r += 1
            teams[name] = rows
        c += 5
    return teams

t2025 = parse_2025(DATI + r"\Rose_fantalba-2025.xlsx")
t2023 = parse_2023(DATI + r"\Rose_fantalba-2023.xlsx")

out = io.open(r"C:\Users\salva\Projects\FantaWarRoom\Excel_Only\_sim4\dump_recompute_v2.txt", "w", encoding="utf-8")
out.write("Squadre 2025 (%d): %s\n" % (len(t2025), list(t2025.keys())))
for k, v in t2025.items():
    out.write("  %s: %d giocatori, costo totale=%.0f\n" % (k, len(v), sum(x[2] or 0 for x in v)))
out.write("Squadre 2023 (%d): %s\n" % (len(t2023), list(t2023.keys())))
for k, v in t2023.items():
    out.write("  %s: %d giocatori, costo totale=%.0f\n" % (k, len(v), sum(x[2] or 0 for x in v)))

# ---------------------------------------------------------------- FVM lookup (Quotazioni)
qwb = openpyxl.load_workbook(DATI + r"\Quotazioni_Fantacalcio_Stagione_2025_26 (2).xlsx", data_only=True)
qws = qwb["Tutti"]
fvm_lookup = {}  # (ruolo, nome_lower) -> fvm
for r in range(3, qws.max_row + 1):
    ruolo = qws.cell(row=r, column=2).value
    nome = qws.cell(row=r, column=4).value
    fvm = qws.cell(row=r, column=12).value
    if ruolo and nome:
        fvm_lookup[(ruolo, str(nome).strip().lower())] = fvm

def match(ruolo, nome):
    key = (ruolo, nome.strip().lower())
    if key in fvm_lookup:
        return fvm_lookup[key]
    return None

# ---------------------------------------------------------------- multiplo prezzo/FVM (SOLO 2025, 8 squadre)
by_role = {"P": [], "D": [], "C": [], "A": []}
matched, total = 0, 0
all_rows = []
for team, rows in t2025.items():
    for ruolo, nome, costo in rows:
        total += 1
        if costo is None:
            continue
        fvm = match(ruolo, nome)
        if fvm is None:
            continue
        matched += 1
        by_role.setdefault(ruolo, []).append((costo, fvm))
        all_rows.append((ruolo, nome, team, costo, fvm, costo / fvm if fvm else None))

out.write("\n--- Match rate 2025 (8 squadre): %d/%d (%.0f%%) ---\n" % (matched, total, matched / total * 100 if total else 0))

out.write("\n--- Multiplo prezzo/FVM per ruolo (2025, 8 squadre) ---\n")
role_names = {"P": "Portieri", "D": "Difensori", "C": "Centrocampisti", "A": "Attaccanti"}
role_stats = {}
for ruolo in ("P", "D", "C", "A"):
    vals = by_role[ruolo]
    n = len(vals)
    costi = [v[0] for v in vals]
    fvms = [v[1] for v in vals]
    multipli = [c / f for c, f in vals if f]
    costo_medio = sum(costi) / n
    fvm_medio = sum(fvms) / n
    mult_medio = sum(multipli) / n
    mult_med = statistics.median(multipli)
    role_stats[ruolo] = dict(n=n, costo_medio=costo_medio, fvm_medio=fvm_medio, mult_medio=mult_medio, mult_mediano=mult_med)
    out.write("  %s: N=%d costo_medio=%.1f fvm_medio=%.1f multiplo_medio=%.2f multiplo_mediano=%.2f\n" %
               (role_names[ruolo], n, costo_medio, fvm_medio, mult_medio, mult_med))

# bidding war (prezzo >> fvm) e sotto-prezzo (prezzo << fvm)
sorted_high = sorted(all_rows, key=lambda x: -(x[5] or 0))
sorted_low = sorted(all_rows, key=lambda x: (x[5] if x[5] is not None else 999))

out.write("\n--- Top 15 CASI PREZZO >> FVM (2025, 8 squadre) ---\n")
for ruolo, nome, team, costo, fvm, mult in sorted_high[:15]:
    out.write("  %s %-22s %-20s costo=%-5.0f fvm=%-5.0f multiplo=%.2f\n" % (ruolo, nome, team, costo, fvm, mult))

out.write("\n--- Top 15 CASI PREZZO << FVM (2025, 8 squadre) ---\n")
for ruolo, nome, team, costo, fvm, mult in sorted_low[:15]:
    if costo < 1:
        continue
    out.write("  %s %-22s %-20s costo=%-5.0f fvm=%-5.0f multiplo=%.2f\n" % (ruolo, nome, team, costo, fvm, mult))

# ---------------------------------------------------------------- % budget per ruolo (2023+2025 combinato, 10 squadre-stagione)
def role_share(all_teams):
    role_sum = {"P": 0, "D": 0, "C": 0, "A": 0}
    total = 0
    for tname, rows in all_teams.items():
        for ruolo, nome, costo in rows:
            if ruolo in role_sum and costo is not None:
                role_sum[ruolo] += costo
                total += costo
    return role_sum, total

out.write("\n--- % budget per ruolo ---\n")
rs2025, tot2025 = role_share(t2025)
out.write("Solo 2025 (8 squadre): totale=%.0f\n" % tot2025)
for r in rs2025:
    out.write("  %s: %.4f\n" % (r, rs2025[r] / tot2025))

rs2023, tot2023 = role_share(t2023)
out.write("Solo 2023 (2 squadre): totale=%.0f\n" % tot2023)
for r in rs2023:
    out.write("  %s: %.4f\n" % (r, rs2023[r] / tot2023))

combined = {}
for k, v in t2025.items():
    combined["2025_" + k] = v
for k, v in t2023.items():
    combined["2023_" + k] = v
rsC, totC = role_share(combined)
out.write("Combinato 2023+2025 (%d squadre-stagione): totale=%.0f\n" % (len(combined), totC))
for r in rsC:
    out.write("  %s: %.4f\n" % (r, rsC[r] / totC))

out.close()
print("Match rate:", matched, "/", total)
print("Role stats:", role_stats)
print("Combined shares (10 squadre-stagione):", {r: round(rsC[r] / totC, 4) for r in rsC})
print("done, dump in dump_recompute_v2.txt")
