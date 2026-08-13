import openpyxl, io

out = io.open(r"C:\Users\salva\Projects\FantaWarRoom\Excel_Only\_sim4\dump_uploaded_check.txt", "w", encoding="utf-8")

paths = {
    "UPLOADED_2023": r"C:\Users\salva\AppData\Roaming\Claude\local-agent-mode-sessions\94c2bcce-5c67-4e3c-bcb2-6a5c6550207a\0d0827d3-ae19-4ee8-9846-c1f409610a73\local_886c534b-88e2-4ce4-9d88-7c954c13cc30\uploads\Rose_fantalba-2023.xlsx",
    "UPLOADED_2025": r"C:\Users\salva\AppData\Roaming\Claude\local-agent-mode-sessions\94c2bcce-5c67-4e3c-bcb2-6a5c6550207a\0d0827d3-ae19-4ee8-9846-c1f409610a73\local_886c534b-88e2-4ce4-9d88-7c954c13cc30\uploads\Rose_fantalba-2025 (1).xlsx",
    "DATI_2023": r"C:\Users\salva\Projects\FantaWarRoom\Excel_Only\Dati\Rose_fantalba-2023.xlsx",
    "DATI_2025": r"C:\Users\salva\Projects\FantaWarRoom\Excel_Only\Dati\Rose_fantalba-2025.xlsx",
}

for label, p in paths.items():
    try:
        wb = openpyxl.load_workbook(p, data_only=True)
    except Exception as e:
        out.write(f"{label}: ERRORE apertura: {e}\n")
        continue
    out.write(f"\n=== {label} ({p}) ===\n")
    out.write("sheets: %s\n" % wb.sheetnames)
    for sn in wb.sheetnames:
        ws = wb[sn]
        out.write(f"  sheet {sn} dim={ws.dimensions} max_col={ws.max_column} max_row={ws.max_row}\n")
        # riga 1 e riga 5 (nomi squadra possono essere su righe diverse)
        for rr in (1, 2, 5, 6):
            vals = [ws.cell(row=rr, column=c).value for c in range(1, min(ws.max_column, 40) + 1)]
            out.write(f"    row{rr}: {vals}\n")
out.close()
print("done")
